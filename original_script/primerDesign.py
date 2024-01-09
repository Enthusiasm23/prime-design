#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Project : TopGen
@Time    : 2022/12/20 10:37
@Author  : lbfeng
@File    :  primerDesign_V3.0.py
@Version : v3.0(20230306)
"""
import os
import io
import re
import sys
import time
import platform
import json
import base64
import datetime
from urllib.parse import urlparse
import concurrent.futures
from email import encoders
from email.mime.base import MIMEBase
from email.utils import formataddr
import tqdm
import yaml
import warnings
import logging
import argparse
import smtplib
import requests
import openpyxl
import email.utils
import pandas as pd
import numpy as np
from email.header import decode_header
from imapclient import IMAPClient
from selenium import webdriver
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from selenium.common import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import ui
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
import primkit

# config
config_path = '/home/ngs/PrimerDesign/script/config.yaml'
if os.path.exists(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        config = yaml.load(f, Loader=yaml.FullLoader)
else:
    with open('config.yaml', 'r', encoding='utf-8') as f:
        config = yaml.load(f, Loader=yaml.FullLoader)

# 测试模式
DEBUG = config['DEBUG']

# 预先订购
advance = config['advance']

# 本地运行
run = config['run']

# 判断系统
system = platform.system()

# 配置日志输出格式
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s - %(message)s')

# 忽略警告
import urllib3
urllib3.disable_warnings()
warnings.simplefilter(action='ignore', category=UserWarning)


def setting():
    options = webdriver.ChromeOptions()
    # detecting system
    if system == 'Linux':
        options.add_argument('--headless')  # 禁止浏览器可视化页面
        options.add_argument('--remote-debugging-port=9222')
        options.add_argument('--user-data-dir=~/.config/google-chrome')
    options.add_argument('--no-sandbox')  # 彻底停用沙箱
    options.add_argument('--disable-infobars')  # 禁用浏览器正在被自动化程序控制的提示
    options.add_argument('--incognito')  # 隐身模式（无痕模式）
    options.add_argument('--ignore-certificate-errors')  # 忽略ssl
    options.add_argument('--disable-gpu')  # 禁用GPU加速
    options.add_argument('--disable-dev-shm-usage')  # 不使用 /dev/shm
    options.add_argument("--disable-notifications")  # 禁用软件光栅化器
    options.add_argument('--disable-software-rasterizer')  # 禁用软件渲染器
    options.add_argument('--enable-logging')  # 详细日志输出
    options.add_experimental_option('excludeSwitches', ['enable-logging'])  # 消除蓝牙适配器错误
    # executable_path = ChromeDriverManager().install()
    # service = ChromeService(executable_path=executable_path)
    # driver = webdriver.Chrome(service=service, options=options)
    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    driver.implicitly_wait(10)
    return driver


def isDriverClosed(driver):
    try:
        driver.get(config['TEST_URL'])
        return False
    except WebDriverException:
        print("webdriver closed by user")
        return True


def is_url(s):
    try:
        result = urlparse(s)
        return all([result.scheme, result.netloc])
    except ValueError:
        return False


def wait_appear_element(driver, locator, by=None, timeout=None):
    if by is None:
        by = By.CSS_SELECTOR
    if timeout is None:
        timeout = 15
    try:
        ui.WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((by, locator)))
        return True
    except TimeoutException:
        return False


def page_load(driver, timeout=None):
    if timeout is None:
        timeout = 10
    times = 0
    while times < timeout:
        time.sleep(1)
        status = driver.execute_script("return document.readyState")
        times = times + 1
        if status == 'complete':
            break


def convert_row_to_string(row):
    row_string = "\t".join(str(x) for x in row)
    return row_string + "\n"


def input_values(driver, element_id, value):
    if wait_appear_element(driver, element_id):
        element = driver.find_element(by=By.CSS_SELECTOR, value=element_id)
        element.click()
        element.clear()
        element.send_keys(value)
    else:
        TimeoutException('ERROR: Failed to get element timeout')
        driver.quit()
        sys.exit(1)


def imitate_web(result_string, url):
    tries = 0
    while tries < config['attempts']:
        try:
            driver = setting()
            try:
                driver.get(url)
            except WebDriverException as e:
                if "net::ERR_NAME_NOT_RESOLVED" in str(e):
                    logging.error('ERROR: network error, please try later.')
                    driver.quit()
                    sys.exit(1)
                else:
                    driver.quit()
                    raise e
            page_load(driver)

            # Regions in BED format
            if wait_appear_element(driver, '#BedInput'):
                input_element = driver.find_element(by=By.CSS_SELECTOR, value='#BedInput')
                input_element.click()
                for i in result_string.split('\r'):
                    for j in i.split('\t')[:-1]:
                        input_element.send_keys(j)
                        time.sleep(0.1)
                        input_element.send_keys(Keys.SHIFT, Keys.SPACE)
                    input_element.send_keys(i.split('\t')[-1])
                    input_element.send_keys(Keys.RETURN)
            else:
                logging.error(
                    f'ERROR: MFEPrimer Page error, positioning element not found.')
                driver.quit()
                sys.exit(1)

            # PrimerMaxSize
            input_values(driver, '#PrimerMaxSize', '25')

            # PrimerMinTm
            input_values(driver, '#PrimerMinTm', '58')

            # PrimerOptTm
            input_values(driver, '#PrimerOptTm', '60')

            # PrimerMaxTm
            input_values(driver, '#PrimerMaxTm', '62')

            # ProdMinSize
            input_values(driver, '#ProdMinSize', '80')

            # ProdMaxSize
            input_values(driver, '#ProdMaxSize', '120')

            # submit
            if wait_appear_element(driver, 'button[type="submit"]'):
                submit_element = driver.find_element(by=By.CSS_SELECTOR, value='button[type="submit"]')
                submit_element.click()
            else:
                logging.error(f'ERROR: MFEPrimer Page error, primer design submission button failed.')
                driver.quit()
                sys.exit(1)

            # Multiplex PCR primer design running
            page_load(driver, timeout=300)
            if wait_appear_element(driver, '//span[contains(text(), "Done")]', by=By.XPATH, timeout=600):
                driver.refresh()
                if wait_appear_element(driver, 'span[class="badge badge-success"]+a'):
                    result_url = driver.find_element(by=By.CSS_SELECTOR,
                                                     value='span[class="badge badge-success"]+a').get_attribute('href')
                    driver.quit()
                    return result_url
                else:
                    logging.error('ERROR: Page error, download link not found.')
                    sys.exit(1)
            else:
                logging.error('ERROR: The network request of primer design exceeds 600 seconds, please try again.')
                driver.quit()
                sys.exit(1)
        except Exception as e:
            if str(e) == "disconnected: unable to connect to renderer":
                tries += 1
            else:
                logging.error(
                    f'ERROR: Network problem or chrome renderer failed to connect, please try again. Error log is {str(e)}')
                raise e


def down_result(down_url, save_name):
    try:
        response = requests.get(down_url, stream=True, timeout=10, verify=False)
        if response.status_code == 200:
            data_size = int(response.headers.get('Content-Length', 0)) / 1024  # change here
            if os.path.exists(save_name):
                logging.warning(f"{save_name} already exist.")
                try:
                    os.remove(save_name)
                except FileNotFoundError:
                    logging.info(f"{save_name} does not exist.")
                logging.info(f"{save_name} deleted and re downloaded.")
            with open(save_name, mode='wb') as f:
                pbar = tqdm.tqdm(total=data_size, unit='KB', desc='downloading')  # and here
                for data in response.iter_content(1024):  # and here
                    f.write(data)
                    pbar.update(len(data) / 1024)  # and here
                pbar.close()
            logging.info(f"{save_name} download complete.")
        else:
            logging.error(f"ERROR: {save_name} download failed, The error code is {response.status_code}")
            sys.exit(1)
    except Exception as e:
        logging.error(f"ERROR: {save_name} download url error, The cause of the error is {e}")
        sys.exit(1)


def select_site(df_source, df_res=None, not_used=None, num=None, driver=None):
    if num is None:
        num = 20
    if df_res is None and not_used is None:
        df_primer = df_source.head(num).copy()
        source_used = df_primer['TemplateID'].to_list()
        not_used = df_source[~df_source['TemplateID'].isin(source_used)]['TemplateID'].to_list()
        result = df_primer[['chrom', 'pos', 'stop']].apply(convert_row_to_string, axis=1)
        result_string = "".join(result)[0:-1]
        return result_string, not_used
    else:
        # 去掉成功的driver
        if driver is not None:
            df_filt = df_res[~(df_res['TemplateID'].isin(driver))].copy()
        else:
            df_filt = df_res.copy()
        # 成功使用的
        sus_used = df_filt['TemplateID'].to_list()
        # 加上未使用的
        sus_used.extend(not_used)
        # next_used = list(set(sus_used))
        # 未使用 source
        df_source = df_source.loc[df_source['TemplateID'].isin(sus_used), :].drop_duplicates('TemplateID',
                                                                                             keep='first').copy()
        # 再选择前20个
        df_future_use = df_source.iloc[:num, :].copy()
        source_used = df_future_use['TemplateID'].to_list()
        not_used = df_source[~df_source['TemplateID'].isin(source_used)]['TemplateID'].to_list()
        result = df_future_use[['chrom', 'pos', 'stop']].apply(convert_row_to_string, axis=1)
        result_string = "".join(result)[0:-1]
        return result_string, not_used


def write_sh_order(df_sample, dataframe, order_path, sampleID):
    os.makedirs(order_path, exist_ok=True)
    try:
        if not run:
            order_template = config['order_template']['sh']
        else:
            order_template = config['local_order_template']['sh']
        wb = openpyxl.load_workbook(order_template)
    except Exception as e:
        logging.error(f'ERROR: {e}')
        sys.exit(1)

    ws_dna = wb['DNA合成订购表']

    def write_to_sheet(data, column, start_row, sequence_prefix=None):
        for i, value in enumerate(data):
            if value.islower():
                value = value.upper()
            if sequence_prefix:
                value = sequence_prefix + value
            ws_dna.cell(start_row + i, column).value = value
            ws_dna.cell(start_row + i, 5).value = 'PAGE'
            ws_dna.cell(start_row + i, 8).value = 1
            ws_dna.cell(start_row + i, 10).value = 2
            ws_dna.cell(start_row + i, 11).value = 'BLG白色标签'
            ws_dna.cell(start_row + i, 12).value = '1.5ml离心管'
            ws_dna.cell(start_row + i, 13).value = 50
            ws_dna.cell(start_row + i, 14).value = 'H+M'
            ws_dna.cell(start_row + i, 15).value = '1*TE 稀释'

    # Write the 'F_id' data to the sheet
    write_to_sheet(df_sample['F_id'], 3, 20)

    # Leave a blank row after writing the 'F_id' data
    ws_dna.cell(20 + len(df_sample['F_id']), 3).value = None

    # Write the 'R_id' data to the sheet
    write_to_sheet(df_sample['R_id'], 3, 21 + len(df_sample['F_id']))

    # Write the 'ForwardPrimer(Fp)' data to the sheet
    write_to_sheet(df_sample['ForwardPrimer(Fp)'], 4, 20, "GTTCAGAGTTCTACAGTCCGACGATCNNWNNW")

    # Leave a blank row after writing the 'ForwardPrimer(Fp)' data
    ws_dna.cell(20 + len(df_sample['ForwardPrimer(Fp)']), 4).value = None

    # Write the 'ReversePrimer(Rp)' data to the sheet
    write_to_sheet(df_sample['ReversePrimer(Rp)'], 4, 21 + len(df_sample['ForwardPrimer(Fp)']),
                   "CTTGGCACCCGAGAATTCCANNWNNW")
    ws_trial = wb['实验用引物对（不需要订购合成）']

    row_index = 1
    # Iterate through the rows of the dataframe
    for r in dataframe.iterrows():
        # Get the row data
        row_data = r[1]
        # Get the cell values
        values = row_data.tolist()
        # Write the data to the sheet
        ws_trial.append(values)
        # Increment the row index
        row_index += 1

    # Save the workbook
    save_file = os.path.join(order_path, '{}_{}_{}.xlsx'.format(sampleID, os.path.basename(
        (config['order_template']['sh']).split('.')[0]), datetime.datetime.now().strftime("%Y%m%d%H%M%S")))
    wb.save(save_file)
    return save_file


def write_hz_order(df_sample, dataframe, order_path, sampleID):
    os.makedirs(order_path, exist_ok=True)
    try:
        if not run:
            order_template = config['order_template']['hz']
        else:
            order_template = config['local_order_template']['hz']
        wb = openpyxl.load_workbook(order_template)
    except Exception as e:
        logging.error(f'ERROR: {e}')
        sys.exit(1)
    ws_dna = wb['DNA合成订购表']

    # Define a function to write the data to the sheet
    def write_to_sheet(data, column, start_row, sequence_prefix=None):
        for i, value in enumerate(data):
            if value.islower():
                value = value.upper()
            if sequence_prefix:
                value = sequence_prefix + value
            ws_dna.cell(start_row + i, column).value = value
            ws_dna.cell(start_row + i, 5).value = 2
            ws_dna.cell(start_row + i, 7).value = 'PAGE'
            ws_dna.cell(start_row + i, 8).value = 2
            ws_dna.cell(start_row + i, 11).value = '1管TE溶解为50uM浓度，1管干粉'

    # Write the 'F_id' data to the sheet
    write_to_sheet(df_sample['F_id'], 2, 20)

    # Leave a blank row after writing the 'F_id' data
    ws_dna.cell(20 + len(df_sample['F_id']), 2).value = None

    # Write the 'R_id' data to the sheet
    write_to_sheet(df_sample['R_id'], 2, 21 + len(df_sample['F_id']))

    # Write the 'ForwardPrimer(Fp)' data to the sheet
    write_to_sheet(df_sample['ForwardPrimer(Fp)'], 3, 20, "GTTCAGAGTTCTACAGTCCGACGATCNNWNNW")

    # Leave a blank row after writing the 'ForwardPrimer(Fp)' data
    ws_dna.cell(20 + len(df_sample['ForwardPrimer(Fp)']), 3).value = None

    # Write the 'ReversePrimer(Rp)' data to the sheet
    write_to_sheet(df_sample['ReversePrimer(Rp)'], 3, 21 + len(df_sample['ForwardPrimer(Fp)']),
                   "CTTGGCACCCGAGAATTCCANNWNNW")

    ws_trial = wb['实验用引物对（不需要订购合成）']

    row_index = 1
    # Iterate through the rows of the dataframe
    for r in dataframe.iterrows():
        # Get the row data
        row_data = r[1]
        # Get the cell values
        values = row_data.tolist()
        # Write the data to the sheet
        ws_trial.append(values)
        # Increment the row index
        row_index += 1

    # Save the workbook
    save_file = os.path.join(order_path, '{}_{}_{}.xlsx'.format(sampleID, os.path.basename(
        (config['order_template']['hz']).split('.')[0]), datetime.datetime.now().strftime("%Y%m%d%H%M%S")))
    wb.save(save_file)
    return save_file


def write_sg_order(df_sample, dataframe, order_path, sampleID):
    os.makedirs(order_path, exist_ok=True)
    try:
        if not run:
            order_template = config['order_template']['sg']
        else:
            order_template = config['local_order_template']['sg']
        wb = openpyxl.load_workbook(order_template)
    except Exception as e:
        logging.error(f'ERROR: {e}')
        sys.exit(1)
    ws_dna = wb['引物合成订购表']

    # Define a function to write the data to the sheet
    def write_to_sheet(data, column, start_row, sequence_prefix=None):
        for i, value in enumerate(data):
            if value.islower():
                value = value.upper()
            if sequence_prefix:
                value = sequence_prefix + value
            ws_dna.cell(start_row + i, column).value = value
            ws_dna.cell(start_row + i, 5).value = 1
            ws_dna.cell(start_row + i, 7).value = 'HAP'
            ws_dna.cell(start_row + i, 8).value = 1
            ws_dna.cell(start_row + i, 12).value = '1管TE溶解为50uM浓度'

    # Write the 'F_id' data to the sheet
    write_to_sheet(df_sample['F_id'], 2, 18)

    # Write the 'R_id' data to the sheet
    write_to_sheet(df_sample['R_id'], 2, 18 + len(df_sample['F_id']))

    # Write the 'ForwardPrimer(Fp)' data to the sheet
    write_to_sheet(df_sample['ForwardPrimer(Fp)'], 3, 18, "GTTCAGAGTTCTACAGTCCGACGATCNNWNNW")

    # Write the 'ReversePrimer(Rp)' data to the sheet
    write_to_sheet(df_sample['ReversePrimer(Rp)'], 3, 18 + len(df_sample['ForwardPrimer(Fp)']),
                   "CTTGGCACCCGAGAATTCCANNWNNW")

    ws_trial = wb['实验用引物对（不需要订购合成）']
    row_index = 1
    # Iterate through the rows of the dataframe
    for r in dataframe.iterrows():
        # Get the row data
        row_data = r[1]
        # Get the cell values
        values = row_data.tolist()
        # Write the data to the sheet
        ws_trial.append(values)
        # Increment the row index
        row_index += 1
    # Save the workbook
    save_file = os.path.join(order_path, '{}_{}_{}.xlsx'.format(sampleID, os.path.basename(
        (config['order_template']['sg']).split('.')[0]), datetime.datetime.now().strftime("%Y%m%d%H%M%S")))
    wb.save(save_file)
    return save_file


def write_dg_order(df_sample, dataframe, order_path, sampleID):
    os.makedirs(order_path, exist_ok=True)
    try:
        if not run:
            order_template = config['order_template']['dg']
        else:
            order_template = config['local_order_template']['dg']
        wb = openpyxl.load_workbook(order_template)
    except Exception as e:
        logging.error(f'ERROR: {e}')
        sys.exit(1)
    ws_dna = wb['订单表格']

    # Define a function to write the data to the sheet
    def write_to_sheet(data, column, start_row, sequence_prefix=None):
        for i, value in enumerate(data):
            if value.islower():
                value = value.upper()
            if sequence_prefix:
                value = sequence_prefix + value
            ws_dna.cell(start_row + i, column).value = value
            # ws_dna.cell(start_row + i, 5).value = 2
            ws_dna.cell(start_row + i, 8).value = 'PAGE'
            ws_dna.cell(start_row + i, 14).value = 2
            ws_dna.cell(start_row + i, 13).value = '1管TE溶解为50uM浓度，1管干粉'

    # Write the 'F_id' data to the sheet
    write_to_sheet(df_sample['F_id'], 5, 16)

    # Write the 'R_id' data to the sheet
    write_to_sheet(df_sample['R_id'], 5, 16 + len(df_sample['F_id']))

    # Write the 'ForwardPrimer(Fp)' data to the sheet
    write_to_sheet(df_sample['ForwardPrimer(Fp)'], 6, 16, "GTTCAGAGTTCTACAGTCCGACGATCNNWNNW")

    # Write the 'ReversePrimer(Rp)' data to the sheet
    write_to_sheet(df_sample['ReversePrimer(Rp)'], 6, 16 + len(df_sample['ForwardPrimer(Fp)']),
                   "CTTGGCACCCGAGAATTCCANNWNNW")

    ws_trial = wb['实验用引物对（不需要订购合成）']
    row_index = 1
    # Iterate through the rows of the dataframe
    for r in dataframe.iterrows():
        # Get the row data
        row_data = r[1]
        # Get the cell values
        values = row_data.tolist()
        # Write the data to the sheet
        ws_trial.append(values)
        # Increment the row index
        row_index += 1
    # Save the workbook
    save_file = os.path.join(order_path, '{}_{}_{}.xlsx'.format(sampleID, os.path.basename(
        (config['order_template']['dg']).split('.')[0]), datetime.datetime.now().strftime("%Y%m%d%H%M%S")))
    wb.save(save_file)
    return save_file


def write_order(file_path, df_source, order_path, mold, snp, skip_review, online=False):
    sampleID = df_source['sampleSn'].iloc[0]
    if 'WX' in sampleID:
        sampleID = sampleID.split('WX')[0]
    elif 'WE' in sampleID:
        sampleID = sampleID.split('WE')[0]
    # 获取结果文件
    if file_path.endswith('.csv'):
        df_sample = pd.DataFrame()
        try:
            df_sample = pd.read_csv(file_path, sep=',', header=3).iloc[0:-1]
        except pd.errors.ParserError:
            with open(file_path, 'r') as f:
                lines = f.readlines()
            # Process the lines
            processed_lines = []
            for i, line in enumerate(lines[4:]):  # Start from the 3rd line (zero-indexed)
                parts = line.split(',')
                if len(parts) > 20:
                    parts[19] = parts[19] + ';' + parts[20]  # Merge the last two fields
                    parts = parts[:20]  # Remove the extra field
                processed_lines.append(','.join(parts))
            # Create a DataFrame from the processed lines
            data = '\n'.join(processed_lines)
            df_sample = pd.read_csv(io.StringIO(data)).iloc[0:-1]
        finally:
            if df_sample.empty:
                logging.error('读取样本 {} iGeneTech 引物设计结果 {} 失败.'.format(sampleID, file_path))
            logging.info('读取样本 {} iGeneTech 引物设计结果 {} 成功.'.format(sampleID, file_path))
    else:
        df_sample = pd.read_excel(file_path, header=3).iloc[0:-1]
    # 设计结果判断：自身位点>=8个,加位点时自身位点小于8个退单, indel和snp时小于8个退单
    if not snp:
        if df_sample.shape[0] < 12:
            if not online:
                if DEBUG:
                    subject = f'【MRD引物设计-测试】{sampleID} 引物结果未通过质控'
                    toaddrs = config['emails']['toaddrs']['log_toaddrs']
                else:
                    subject = f'【MRD引物设计】{sampleID} 引物结果未通过质控'
                    toaddrs = config['emails']['toaddrs']['qc_toaddrs']

                emit(
                    message=f'MRD样本ID：{sampleID}\n\n质控结果：引物结果数量为 {df_sample.shape[0]}, 当样本引物设计完成时数量小于12, 需要审核人员审核退单!\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                    toaddrs=toaddrs,
                    cc=config['emails']['toaddrs']['log_cc'],
                    subject=subject,
                    annex_path=file_path
                )

                logging.error(f'ERROR: MRD样本ID：{sampleID}, 已发邮件至审核人员退单!')
                sys.exit(0)
            else:
                logging.error(
                    f'ERROR: MRD样本ID：{sampleID}, 引物结果未通过质控, 质控结果: 引物结果数量为 {df_sample.shape[0]}, 当样本引物设计完成时数量小于12, 需要审核人员审核退单!')
                sys.exit(0)

        elif df_sample.shape[0] >= 12 and 'hots' in df_source.columns:
            df_hots = pd.merge(df_sample[['TemplateID']].copy(), df_source[['TemplateID', 'hots']].copy(),
                               on='TemplateID').drop_duplicates('TemplateID', keep='first')
            if df_hots[df_hots['hots'] == 0].shape[0] < 8:
                if not online:
                    if DEBUG:
                        subject = f'【MRD引物设计-测试】{sampleID} 引物结果未通过质控'
                        toaddrs = config['emails']['toaddrs']['log_toaddrs']
                    else:
                        subject = f'【MRD引物设计】{sampleID} 引物结果未通过质控'
                        toaddrs = config['emails']['toaddrs']['qc_toaddrs']
                    if not check_emil_sent(subject):
                        emit(
                            message=f'MRD样本ID：{sampleID}\n\n质控结果：当样本自身位点 + 热点进行引物设计完成时，自身位点数量为{df_hots[df_hots["hots"] == 0].shape[0]}，当自身位点数量小于8，需要审核人员审核退单！\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                            toaddrs=toaddrs,
                            cc=config['emails']['toaddrs']['log_cc'],
                            subject=subject,
                            annex_path=file_path
                        )
                    logging.info(f'ERROR: 样本SampleSn为:{sampleID}, 已发邮件至审核人员退单!')
                    sys.exit(0)
                else:
                    logging.info(
                        f'ERROR: MRD样本ID: {sampleID}, 质控结果: 当样本自身位点 + 热点进行引物设计完成时, 自身位点数量为: {df_hots[df_hots["hots"] == 0].shape[0]}, 当自身位点数量小于8, 需要审核人员审核退单!')
                    sys.exit(0)

    df_source = df_source.drop_duplicates().copy()
    if 'hots' in df_source.columns:
        df_source_filt = df_source[
            ['sampleSn', 'chrom', 'Start_Position', 'ref', 'alt', 'gene', 'vaf', 'depth', 'TemplateID', 'hots', 'cHGVS',
             'pHGVS']].copy()
        df_source_filt.rename(columns={'Start_Position': 'pos'}, inplace=True)
    else:
        df_source_filt = df_source[
            ['sampleSn', 'chrom', 'pos', 'ref', 'alt', 'gene', 'vaf', 'depth', 'TemplateID', 'cHGVS', 'pHGVS']].copy()
        df_source_filt['pos'] = df_source_filt['pos'] + 1
    # 合并源文件与结果文件
    if 'hots' in df_source_filt.columns:
        df_sample = pd.merge(df_sample, df_source_filt, on='TemplateID').drop_duplicates('TemplateID', keep='first')
    else:
        df_sample = pd.merge(df_sample, df_source_filt, on='TemplateID')
    df_sample['F_id'] = df_sample.apply(
        lambda row: 'P' + row['sampleSn'].split('NGS')[-1].split('W')[0] + '-' + '{:02d}'.format(
            row.name + 1) + 'F', axis=1)
    df_sample['R_id'] = df_sample.apply(
        lambda row: 'P' + row['sampleSn'].split('NGS')[-1].split('W')[0] + '-' + '{:02d}'.format(
            row.name + 1) + 'R', axis=1)
    df_sample['primerID'] = df_sample.apply(
        lambda row: 'P' + row['sampleSn'].split('NGS')[-1].split('W')[0] + '-' + '{:02d}'.format(
            row.name + 1), axis=1)
    df_sample['ForwardPrimer(Fp)'] = df_sample['ForwardPrimer(Fp)'].apply(lambda x: x.upper())
    df_sample['ReversePrimer(Rp)'] = df_sample['ReversePrimer(Rp)'].apply(lambda x: x.upper())
    # 添加Selected
    df_sample['Selected'] = 1
    # 添加通用序列大小
    df_sample['GnlAmpSize (bp)'] = df_sample['AmpSize(bp)'] + 140
    df_sample['failed_reason'] = ''
    dataframe = df_sample[
        ['sampleSn', 'chrom', 'pos', 'ref', 'alt', 'gene', 'vaf', 'depth', 'primerID', 'Selected', 'failed_reason',
         'ID', 'TemplateID', 'ForwardPrimer(Fp)', 'ReversePrimer(Rp)', 'FpTm', 'RpTm', 'FpSize', 'RpSize', 'FpGC(%)',
         'RpGC(%)', 'FpPos', 'RpPos', 'AmpSize(bp)', 'GnlAmpSize (bp)', 'AmpGC', 'AmpPos', 'Penalty', 'Chr', 'Start',
         'End', 'cHGVS', 'pHGVS', 'Note']].copy()

    # 引物设计结果文件路径
    primer_result = None
    if mold == 'sh':
        primer_result = write_sh_order(df_sample, dataframe, order_path, sampleID)
        logging.info('Primer design Shanghai order template writing completed.')
    elif mold == 'hz':
        primer_result = write_hz_order(df_sample, dataframe, order_path, sampleID)
        logging.info('Primer design Huzhou order template writing completed.')
    elif mold == 'sg':
        primer_result = write_sg_order(df_sample, dataframe, order_path, sampleID)
        logging.info('Primer design ShengGong order template writing completed.')
    elif mold == 'dg':
        primer_result = write_dg_order(df_sample, dataframe, order_path, sampleID)
        logging.info('Primer design ShengGong order template writing completed.')

    # if 55249071 not in dataframe['pos'].to_list():
    #     subject = f'{sampleID} 无T90M引物'
    #     toaddrs = config['emails']['toaddrs']['log_toaddrs']
    #     cc = config['emails']['toaddrs']['log_cc']
    #     emit(
    #         message=f'注：该样本无T90M引物。\n\n引物结果：{os.path.basename(primer_result)}（见附件）\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
    #         toaddrs=toaddrs,
    #         cc=cc,
    #         subject=subject,
    #         annex_path=primer_result,
    #     )

    if online:
        logging.info(f'Primer result path is 172.16.10.9: {primer_result}')
        print(primer_result)
    else:
        # 检查cms系统审核状态, 早中晚, 3天, 发送邮箱
        check_send_mail(sampleID, primer_result, skip_review)


def check_send_mail_old(sampleSn, primer_result):
    # 预先订购
    if advance:
        if DEBUG:
            subject = f'【MRD引物设计-测试】{sampleSn} 引物合成订购'
            toaddrs = config['emails']['toaddrs']['log_toaddrs']
            cc = config['emails']['toaddrs']['log_cc']
        else:
            subject = f'【MRD引物设计】{sampleSn} 引物合成订购'
            toaddrs = config['emails']['toaddrs']['order_toaddrs']
            cc = config['emails']['toaddrs']['cc']
        if not check_emil_sent(subject):
            emit(
                message=f'MRD样本ID：{sampleSn}\n\nCMS审核结果：非正常审核或接口问题（预先订购）\n\nMRD引物结果：{os.path.basename(primer_result)}（见附件）\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                toaddrs=toaddrs,
                cc=cc,
                subject=subject,
                annex_path=primer_result,
            )
    else:
        # 设置早中晚3次检查状态
        morning_time = datetime.time(8, 0)
        noon_time = datetime.time(12, 0)
        night_time = datetime.time(17, 0)
        start_time = datetime.datetime.now()
        is_first_check = True
        days_since_last_email = 0
        check_frequency = datetime.timedelta(days=1)
        max_days_to_check = 15
        while True:
            sample_local = determine_sample_location(sampleSn)
            if sample_local == 'OLD':
                audit_status = get_wes_check_status_old(sampleSn)
                if audit_status:
                    audit_status = 'YWC'
                    status = '已完成'
                else:
                    audit_status = 'DSH'
                    status = '待审核'
            elif sample_local == 'NEW':
                # 千翼接口
                audit_status = get_wes_check_status(sampleSn)
                status_dict = {'YWC': '已完成', 'DSY': '待收样', 'JCZZ': '检测终止', 'BYZ': '补样中', 'YSH': '已审核', 'WTG': '审核未通过',
                               'DSH': '待审核', 'BHG': '不合格', 'FJZ': '复检中', 'ZTJC': '暂停检测', 'JCZ': '检测中', 'YSY': '已收样',
                               'YSC': '已送出', 'DTJ': '待提交', 'TYZ': '退样中', 'YTY': '已退样', 'YZF': '已作废', 'YCY': '已采样'}
                status = status_dict[audit_status]

                # TopGen(上传报告接口)
                # audit_status = get_wes_check_status_new(sampleSn)
                # if audit_status:
                #     audit_status = 'YWC'
                #     status = '已完成'
                # else:
                #     audit_status = 'DSH'
                #     status = '待审核'
            else:
                logging.error(
                    f'ERROR: The sample {sampleSn} does not exist in the small wide CMS system, nor does it exist in the thousand wing CMS system')
                sys.exit(1)

            if audit_status in ['YWC', 'YSH']:
                if DEBUG:
                    subject = f'【MRD引物设计-测试】{sampleSn} 引物合成订购'
                    toaddrs = config['emails']['toaddrs']['log_toaddrs']
                    cc = config['emails']['toaddrs']['log_cc']
                else:
                    subject = f'【MRD引物设计】{sampleSn} 引物合成订购'
                    toaddrs = config['emails']['toaddrs']['order_toaddrs']
                    cc = config['emails']['toaddrs']['cc']
                if not check_emil_sent(subject):
                    emit(
                        message=f'MRD样本ID：{sampleSn}\n\nCMS审核结果：已通过\n\nMRD引物结果：{os.path.basename(primer_result)}（见附件）\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                        toaddrs=toaddrs,
                        cc=cc,
                        subject=subject,
                        annex_path=primer_result,
                    )
                break
            else:
                if is_first_check:
                    if DEBUG:
                        subject = f'【MRD引物设计-测试】{sampleSn} 持续检测样本审核状态'
                    else:
                        subject = f'【MRD引物设计】{sampleSn} 持续检测样本审核状态'
                    emit(
                        message=f'MRD样本ID：{sampleSn}\n\nCMS审核结果：{status}\n\n正在持续检测CMS系统样本审核状态！\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                        subject=subject, annex_path=primer_result)
                    is_first_check = False
                current_time = datetime.datetime.now().time()
                logging.info(f'The current time is {current_time}, Waiting {(datetime.datetime.now() - start_time)}')
                if current_time < morning_time or current_time >= night_time:
                    if current_time >= night_time:
                        next_day = datetime.datetime.now() + datetime.timedelta(days=1)
                        time_to_wait = datetime.datetime.combine(next_day, morning_time) - datetime.datetime.now()
                    else:
                        time_to_wait = datetime.datetime.combine(datetime.datetime.now(),
                                                                 morning_time) - datetime.datetime.now()
                    time.sleep(time_to_wait.total_seconds())
                elif current_time < noon_time:
                    time_to_wait = datetime.datetime.combine(datetime.datetime.now(),
                                                             noon_time) - datetime.datetime.now()
                    time.sleep(time_to_wait.total_seconds())
                elif current_time < night_time:
                    time_to_wait = datetime.datetime.combine(datetime.datetime.now(),
                                                             night_time) - datetime.datetime.now()
                    time.sleep(time_to_wait.total_seconds())
                current_time = datetime.datetime.now()
                time_diff = current_time - start_time
                logging.info(f'The current time is {current_time}, Waiting {time_diff}')
                if time_diff >= check_frequency:
                    days_since_last_email += 1
                    if days_since_last_email % 3 == 0 and days_since_last_email > 0:
                        current_cycle = days_since_last_email // 3
                        if current_cycle >= max_days_to_check // 3:
                            if DEBUG:
                                subject = f'【MRD引物设计-测试】{sampleSn} 样本审核状态超过半个月未更新提醒'
                                toaddrs = config['emails']['toaddrs']['log_toaddrs']
                            else:
                                subject = f'【MRD引物设计】{sampleSn} 样本审核状态超过半个月未更新提醒'
                                toaddrs = config['emails']['toaddrs']['qc_toaddrs']
                            if not check_emil_sent(subject):
                                emit(
                                    message=f'MRD样本ID：{sampleSn}\n\nCMS审核结果：检测到已经超过半个月未通过审核，请审核人员检查并更新状态！\n\n注意：该样本审核状态最后一次检测！\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                                    subject=subject,
                                    toaddrs=toaddrs,
                                    cc=config['emails']['toaddrs']['log_cc'],
                                )
                            logging.info(
                                f'The program has been running for more than {max_days_to_check} days. Exiting program.')
                            break
                        elif days_since_last_email % 3 == 0 and days_since_last_email > 0:
                            if DEBUG:
                                subject = f'【MRD引物设计-测试】{sampleSn} 样本审核状态超过{current_cycle * 3} 天未更新提醒'
                                toaddrs = config['emails']['toaddrs']['log_toaddrs']
                            else:
                                subject = f'【MRD引物设计】{sampleSn} 样本审核状态超过{current_cycle * 3} 天未更新提醒'
                                toaddrs = config['emails']['toaddrs']['qc_toaddrs']
                            if not check_emil_sent(subject):
                                emit(
                                    message=f'MRD样本ID：{sampleSn}\n\nCMS审核结果：检测到已经超过{current_cycle * 3} 天未通过审核，请审核人员检查并更新状态！\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                                    subject=subject,
                                    toaddrs=toaddrs,
                                    cc=config['emails']['toaddrs']['log_cc'],
                                )
                            check_frequency = datetime.timedelta(days=3)
                        else:
                            check_frequency = datetime.timedelta(days=1)
                else:
                    time_to_wait = check_frequency - time_diff
                    time.sleep(time_to_wait.total_seconds())


def generate_testing_periods(cycle, max_cycle):
    periods = [cycle]
    for i in range(cycle * 2, max_cycle + 1, cycle):
        periods.append(i)
    if periods[-1] != max_cycle:
        periods.append(max_cycle)
    return periods


def check_send_mail(sampleSn, primer_result, skip_review):
    # 预先订购
    if advance:
        if DEBUG:
            subject = f'【MRD引物设计-测试】{sampleSn} 引物合成订购'
            toaddrs = config['emails']['toaddrs']['log_toaddrs']
            cc = config['emails']['toaddrs']['log_cc']
        else:
            subject = f'【MRD引物设计】{sampleSn} 引物合成订购'
            toaddrs = config['emails']['toaddrs']['order_toaddrs']
            cc = config['emails']['toaddrs']['cc']
        if not check_emil_sent(subject):
            emit(
                message=f'MRD样本ID：{sampleSn}\n\nCMS审核结果：非正常审核（预先订购）\n\nMRD引物结果：{os.path.basename(primer_result)}（见附件）\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                toaddrs=toaddrs,
                cc=cc,
                subject=subject,
                annex_path=primer_result,
            )
    elif skip_review:
        if DEBUG:
            subject = f'【MRD引物设计-测试】{sampleSn} 引物合成订购 ( 预先订购 | 位点追加 | 项目测试 | 科研项目)'
            toaddrs = config['emails']['toaddrs']['log_toaddrs']
            cc = config['emails']['toaddrs']['log_cc']
        else:
            subject = f'【MRD引物设计】{sampleSn} 引物合成订购 ( 预先订购 | 位点追加 | 项目测试 | 科研项目)'
            toaddrs = config['emails']['toaddrs']['order_toaddrs']
            cc = config['emails']['toaddrs']['cc']
        emit(
            message=f'注：该引物合成用于(预先订购 | 位点追加 | 项目测试 | 科研项目)其中之一。\n\n引物结果：{os.path.basename(primer_result)}（见附件）\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
            toaddrs=toaddrs,
            cc=cc,
            subject=subject,
            annex_path=primer_result,
        )
    else:
        program_time = datetime.datetime.now()
        start_time = datetime.datetime.now()
        check_interval_minutes = int(config['check_interval_minutes'])    # 检测时间
        check_frequency = datetime.timedelta(minutes=check_interval_minutes)    # min转换s
        email_cycle = int(config['email_interval_days'])     # 邮件预警周期
        max_days_to_check = int(config['max_interval_days'])     # 最大检测周期
        email_days = generate_testing_periods(email_cycle, max_days_to_check)   # 周期内天数
        is_first_check = True   # 持续监测预警
        last_email_sent = None  # 发送邮件标志

        while True:
            sample_local = determine_sample_location(sampleSn)
            if sample_local == 'OLD':
                # 小阔接口
                status = get_sample_status_old(sampleSn)
                status_dict = {'YCY': '已采样', 'YSC': '已送出', 'YSY': '已收样', 'JCZ': '检测中', 'FJZ': '复检中', 'YWC': '已完成', 'JCZZ': '检测终止', 'BHG': '不合格', 'BGDSH': '报告待审核', 'BGWTG': '报告审核未通过', 'BGYSH': '报告已审核', 'BYZ': '补样中', 'ZTJC': '暂停检测'}
                audit_status = next((key for key, value in status_dict.items() if value == status), None)
            elif sample_local == 'NEW':
                # 千翼接口
                audit_status = get_wes_check_status(sampleSn)
                status_dict = {'YWC': '已完成', 'DSY': '待收样', 'JCZZ': '检测终止', 'BYZ': '补样中', 'YSH': '已审核', 'WTG': '审核未通过', 'DSH': '待审核', 'BHG': '不合格', 'FJZ': '复检中', 'ZTJC': '暂停检测', 'JCZ': '检测中', 'YSY': '已收样', 'YSC': '已送出', 'DTJ': '待提交', 'TYZ': '退样中', 'YTY': '已退样', 'YZF': '已作废', 'YCY': '已采样'}
                status = status_dict[audit_status]
            else:
                logging.error(
                    f'ERROR: The sample {sampleSn} does not exist in the small wide CMS system, nor does it exist in the thousand wing CMS system')
                sys.exit(1)

            if audit_status in ['JCZZ', 'ZTJC']:
                if DEBUG:
                    subject = f'【MRD引物设计-测试】{sampleSn} 样本检测异常提醒'
                else:
                    subject = f'【MRD引物设计】{sampleSn} 样本检测异常提醒'
                emit(
                    message=f'MRD样本ID：{sampleSn}\n\nCMS审核结果：{status}\n\n退出引物设计\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                    subject=subject, annex_path=primer_result)
                break
            # 审核通过
            elif audit_status in ['YWC', 'YSH', 'BGYSH']:
                if DEBUG:
                    subject = f'【MRD引物设计-测试】{sampleSn} 引物合成订购'
                    toaddrs = config['emails']['toaddrs']['log_toaddrs']
                    cc = config['emails']['toaddrs']['log_cc']
                else:
                    subject = f'【MRD引物设计】{sampleSn} 引物合成订购'
                    toaddrs = config['emails']['toaddrs']['order_toaddrs']
                    cc = config['emails']['toaddrs']['cc']
                if not check_emil_sent(subject):
                    emit(
                        message=f'MRD样本ID：{sampleSn}\n\nCMS审核结果：已通过\n\nMRD引物结果：{os.path.basename(primer_result)}（见附件）\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                        toaddrs=toaddrs,
                        cc=cc,
                        subject=subject,
                        annex_path=primer_result,
                    )
                break
            else:
                # 第一次检测未审核 发邮件提醒
                if is_first_check:
                    if DEBUG:
                        subject = f'【MRD引物设计-测试】{sampleSn} 持续检测样本审核状态'
                    else:
                        subject = f'【MRD引物设计】{sampleSn} 持续检测样本审核状态'
                    emit(
                        message=f'MRD样本ID：{sampleSn}\n\nCMS审核结果：{status}\n\n正在持续检测CMS系统样本审核状态！\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                        subject=subject, annex_path=primer_result)
                    is_first_check = False

                # 根据自定义时间间隔等待下一次检测
                current_time = datetime.datetime.now()

                # 等待时间
                time_to_wait = check_frequency - (current_time - start_time)
                if time_to_wait.total_seconds() > 0:
                    time.sleep(time_to_wait.total_seconds())

                # 等待后继续算时间间隔
                current_time = datetime.datetime.now()
                time_diff = current_time - start_time
                logging.info(f'The current time is {current_time}, Waiting {time_diff}')

                # 若时间间隔已经大于检测间隔时间
                if time_diff >= check_frequency:
                    # 计算过去的天数
                    days_since_last = (current_time - program_time).days
                    days_since_last_email = int(days_since_last)
                    program_time_formatted = program_time.strftime('%Y-%m-%d %H:%M:%S:%f')
                    current_time_formatted = current_time.strftime('%Y-%m-%d %H:%M:%S:%f')
                    if days_since_last_email in email_days:
                        if days_since_last_email < max_days_to_check:
                            if last_email_sent is None or (current_time - last_email_sent).days >= email_cycle:
                                if DEBUG:
                                    subject = f'【MRD引物设计-测试】{sampleSn} 样本审核状态超过{days_since_last_email} 天未更新提醒'
                                    toaddrs = config['emails']['toaddrs']['log_toaddrs']
                                else:
                                    subject = f'【MRD引物设计】{sampleSn} 样本审核状态超过{days_since_last_email} 天未更新提醒'
                                    toaddrs = config['emails']['toaddrs']['qc_toaddrs']
                                if not check_emil_sent(subject):
                                    emit(
                                        message=f'MRD样本ID：{sampleSn}\n\nCMS审核结果：检测到已经超过{days_since_last_email} 天未通过审核，请审核人员检查并更新状态！\n\n检测时间：{program_time_formatted} —— {current_time_formatted}\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                                        subject=subject,
                                        toaddrs=toaddrs,
                                        cc=config['emails']['toaddrs']['log_cc'],
                                    )
                                    last_email_sent = current_time
                        else:
                            if last_email_sent is None or (current_time - last_email_sent).days >= email_cycle:
                                if DEBUG:
                                    subject = f'【MRD引物设计-测试】{sampleSn} 样本审核状态超过半个月未更新提醒'
                                    toaddrs = config['emails']['toaddrs']['log_toaddrs']
                                else:
                                    subject = f'【MRD引物设计】{sampleSn} 样本审核状态超过半个月未更新提醒'
                                    toaddrs = config['emails']['toaddrs']['qc_toaddrs']
                                if not check_emil_sent(subject):
                                    emit(
                                        message=f'MRD样本ID：{sampleSn}\n\nCMS审核结果：检测到已经超过半个月未通过审核，请审核人员检查并更新状态！\n\n检测时间：{program_time_formatted} —— {current_time_formatted}\n\n注意：该样本审核状态最后一次检测！\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                                        subject=subject,
                                        toaddrs=toaddrs,
                                        cc=config['emails']['toaddrs']['log_cc'],
                                    )
                                    logging.info(
                                        f'The program has been running for more than {max_days_to_check} days. Exiting program.')
                                break
                    # 重置 start_time
                    start_time = datetime.datetime.now()
                else:
                    time_to_wait = check_frequency - time_diff
                    if time_to_wait.total_seconds() > 0:
                        time.sleep(time_to_wait.total_seconds())


def get_cms_accessToken():
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'}
    postUrl = config['CMS_URL']['accessToken']['post_url']
    postData = config['CMS_URL']['accessToken']['post_data']
    try:
        response = requests.post(postUrl, params=postData, headers=headers)
        if response.status_code == 200:
            accessToken = json.loads(response.text)['data']['accessToken']
            logging.info(f'AccessToken obtained successfully, accessToken: {accessToken}')
            return accessToken
        else:
            logging.error('ERROR: response.status_code is not equal to 200.')
            sys.exit(1)
    except Exception as e:
        logging.error(f'ERROR: Error getting sample audit status of cms system. The specific reason is {e}')
        sys.exit(1)


def get_cms_api_token():
    headers = config['header']
    postUrl = config['CMS_URL']['apiToken']['post_url']
    postData = config['CMS_URL']['apiToken']['post_data']
    try:
        response = requests.post(postUrl, data=postData, headers=headers)
        if response.status_code == 200:
            text = response.text
            try:
                token = json.loads(text)["result"]["token"]
            except KeyError:
                token = json.loads(text)["data"]["token"]
            logging.info(f'Token obtained successfully, Token: {token}')
            return token
        else:
            logging.error('ERROR: response.status_code is not equal to 200.')
            sys.exit(1)
    except Exception as e:
        logging.error(f'ERROR: Error getting sample audit status of cms system. The specific reason is {e}')
        sys.exit(1)


def determine_sample_location(sampleSn):
    """
    sampleSn: sample ID
    return: NEW: 千翼CMS系统, OLD: 小阔CMS系统, NONE: CMS不存在
    """
    old_cms = config['API_CMS']['old_cms']['post_url']
    old_search = config['API_CMS']['old_search']['get_url']
    new_cms = config['API_CMS']['new_cms']['detail_url']

    old_up = config['API_CMS']['old_cms']['post_data']
    old_header = config['header']
    old_token = json.loads(requests.post(old_cms, params=old_up, headers=old_header).text)['data']['accessToken']
    old_payload = {'accessToken': old_token, "search[sampleSn][value]": sampleSn, "search[sampleSn][query]": "eq"}
    old_response = requests.get(old_search, params=old_payload, headers=old_header)
    if old_response.status_code == 200:
        old_data = json.loads(old_response.text)['data']
    else:
        logging.error('ERROR: old system response.status_code is not equal to 200.')
        sys.exit(1)

    new_payload = json.dumps({"fybh": sampleSn})
    new_header = config['API_CMS']['new_cms']['new_header']
    new_response = requests.post(new_cms, data=new_payload, headers=new_header, timeout=30)
    if new_response.status_code == 200:
        new_data = json.loads(new_response.text)
        if new_data['errorCode'] == '0' or new_data['msg'] == 'OK':
            new_data = new_data['data']
        elif new_data['errorCode'] == '400' or new_data['msg'] == '样本数据为空':
            new_data = []
        else:
            logging.error(
                'ERROR: new system errorCode is {}. errorMsg is {}.'.format(new_data['errorCode'],
                                                                            new_data['msg']))
            sys.exit(1)
    else:
        logging.error('ERROR: new system response.status_code is not equal to 200.')
        sys.exit(1)

    if old_data and not new_data:
        return 'OLD'
    elif not old_data and new_data:
        return 'NEW'
    else:
        return 'NONE'


def get_project_itemName_old(sampleSn):
    accessToken = get_cms_accessToken()
    sampleInfo_url = config['CMS_URL']['sampleInfo']['get_url']
    payload = {'accessToken': accessToken, "search[sampleSn][value]": sampleSn, "search[sampleSn][query]": "eq"}
    try:
        result = requests.get(sampleInfo_url, params=payload)
        dicts = json.loads(result.text)
        if len(dicts["data"]) > 0:
            project_id = dicts["data"][0]["itemName"]
            logging.info(f'Successfully obtained project ID, project ID: {project_id}')
            return project_id
        else:
            logging.warning('ERROR: Sample ID does not exist in cms system!')
            return 'OTHER'
    except Exception as e:
        logging.error(
            f'ERROR: An error occurred while getting the item ID of the sample for the cms system. The specific reason is {e}')
        sys.exit(1)


def get_project_itemName(sampleSn):
    payload = {"fybh": sampleSn}
    post_data = json.dumps(payload)
    post_url = config['API_CMS']['new_cms']['detail_url']
    header = config['API_CMS']['new_cms']['new_header']
    response = requests.post(post_url, data=post_data, headers=header, timeout=30)
    if response.status_code == 200:
        sample_info = json.loads(response.text)
        if sample_info['errorCode'] == '0':
            return sample_info['data']['DD'][0]['XMMC']
        else:
            logging.error(
                'ERROR: errorCode is {}. errorMsg is {}.'.format(sample_info['errorCode'], sample_info['msg']))
            sys.exit(1)
    else:
        logging.error('ERROR: response.status_code is not equal to 200.')
        sys.exit(1)


def get_project_type_old(sampleSn):
    project_id = get_project_itemName_old(sampleSn)
    project_id = re.search(r"^(.{7})", project_id).group(1)
    MRD_detection = config['MRD_ID']
    if project_id in MRD_detection:
        return "MRD"
    elif project_id == "OTHER":
        return "OTHER"
    else:
        return "OTHER"


def get_project_type(sampleSn):
    itemName = get_project_itemName(sampleSn)
    if '迈锐达' in itemName or 'MRD' in itemName:
        return "MRD"
    else:
        return "OTHER"


def get_wes_check_status_old(sampleSn):
    token = get_cms_api_token()
    session = requests.session()
    header = {'User-Agent': config['header']['User-Agent'], "Access-Token": token}
    postUrl = config['CMS_URL']['check_audit']['post_url']
    postData = {'sample_no': sampleSn}
    try:
        response = session.post(postUrl, params=postData, headers=header)
        if response.status_code == 200:
            result = json.loads(response.text)
            wes_check_status = result["data"]["sample"]["auth_status"]
            logging.info(f'获取样本{sampleSn}审核状态为: {wes_check_status}.')
            return wes_check_status
        else:
            logging.error("ERROR: 获取样本 {} 审核状态失败, 响应代码为 {}.".format(sampleSn, response.status_code))
            sys.exit(1)
    except Exception as e:
        logging.error("ERROR: 获取样本 {} 审核状态失败, 错误为: {}.".format(sampleSn, e))
        sys.exit(1)


def get_sample_status_old(sampleSn):
    accessToken = get_cms_accessToken()
    sampleInfo_url = config['CMS_URL']['sampleInfo']['get_url']
    payload = {'accessToken': accessToken, "search[sampleSn][value]": sampleSn, "search[sampleSn][query]": "eq"}
    try:
        result = requests.get(sampleInfo_url, params=payload)
        dicts = json.loads(result.text)
        if len(dicts["data"]) > 0:
            sampleStatus = dicts["data"][0]["sampleStatusShow"]
            logging.info(f'Successfully obtained sampleStatus, sampleStatus: {sampleStatus}')
            return sampleStatus
        else:
            logging.warning('ERROR: Sample ID does not exist in cms system!')
            return 'OTHER'
    except Exception as e:
        logging.error(
            f'ERROR: An error occurred while getting the item ID of the sample for the cms system. The specific reason is {e}')
        sys.exit(1)


def get_wes_check_status(sampleSn):
    payload = {"fybh": sampleSn}
    post_data = json.dumps(payload)
    post_url = config['API_CMS']['new_cms']['detail_url']
    header = config['API_CMS']['new_cms']['new_header']
    response = requests.post(post_url, data=post_data, headers=header, timeout=30)
    if response.status_code == 200:
        sampleSn_info = json.loads(response.text)
        if sampleSn_info['errorCode'] == '0':
            sample_status = sampleSn_info['data']['YBFY']['YBZT']
            return sample_status
        else:
            logging.error(
                'ERROR: errorCode is {}. errorMsg is {}.'.format(sampleSn_info['errorCode'], sampleSn_info['msg']))
            sys.exit(1)
    else:
        logging.error('ERROR: response.status_code is not equal to 200.')
        sys.exit(1)


def get_wes_check_status_new(sampleSn):
    """
    通过接口的update_time字段判断样品是否通过审核，以后可能会更改,使用其他字段或者接口
    返回：True or False
    """
    get_url = config['API_CMS']['audit_cms']['get_url'] + sampleSn
    header = config['header']
    response = requests.get(get_url, headers=header, timeout=30)
    if response.status_code == 200:
        content = json.loads(response.text)
        if content["data"]["sample"] is None or content["data"]["sample"] == "":
            return False
        if "upload_time" in content["data"]["sample"].keys() and content["data"]["sample"]["upload_time"] != "":
            upload_time = content["data"]["sample"]["upload_time"]
            logging.info(f'The report was uploaded on {upload_time}.')
            return True
        else:
            return False
    else:
        logging.error(f'Error: {get_url} response.status_code is not equal to 200.')
        sys.exit(1)


def emit(message, toaddrs=None, cc=None, subject=None, annex_path=None):
    try:
        mail_host: str = config['emails']['setting']['mail_hot']
        mail_port: int = config['emails']['setting']['mail_port']
        username: str = config['emails']['setting']['username']
        password: str = config['emails']['setting']['password']
        from_addr: str = config['emails']['setting']['from_addr']
        if toaddrs is None:
            toaddrs: list[str] = config['emails']['setting']['send_addr']['sender']
        if subject is None:
            subject: str = 'MRD引物设计选点未通过质控'
        timeout: float = 60
        if not mail_port:
            mail_port = smtplib.SMTP_SSL_PORT
        smtp = smtplib.SMTP_SSL(mail_host, mail_port, timeout=timeout)
        msg = MIMEMultipart()
        # 发件人 发件人名称
        sender_name = config['emails']['setting']['send_addr']['sender_name']
        msg['From'] = formataddr(pair=(sender_name, from_addr))
        # 收件人
        msg['To'] = ','.join(toaddrs)
        # 抄送人
        if cc is not None:
            msg['Cc'] = ','.join(cc)
        # 主题
        msg['Subject'] = subject
        # 日期
        msg['Date'] = str(email.utils.localtime())
        # 正文
        text = MIMEText(_text=message, _charset='utf-8')
        msg['Content-Type'] = 'text/html'
        msg.attach(text)
        # 附件
        if annex_path is not None:
            content_type = 'application/octet-stream'
            maintype, subtype = content_type.split('/', 1)
            # 读入文件内容并格式化
            data = open(annex_path, 'rb')
            file_msg = MIMEBase(maintype, subtype)
            file_msg.set_payload(data.read())
            data.close()
            encoders.encode_base64(file_msg)
            # 设置附件头
            basename = os.path.basename(annex_path)
            # 解决中文附件名乱码问题
            file_msg.add_header('Content-Disposition', 'attachment', filename=('gbk', '', basename))
            msg.attach(file_msg)
        # 登录
        smtp.login(username, password)
        # 发送
        smtp.send_message(msg)
        smtp.quit()
        if '未通过质控' in subject:
            logging.info('MRD引物设计质控提醒邮件发送成功!')
        else:
            logging.info('邮箱发送成功!')
    except Exception as esg:
        logging.error(f'ERROR: 发送邮件失败, 原因为: {esg}')
        sys.exit(1)


def check_emil_sent(param_subject, use_multithread=False):
    def get_available_threads():
        import os
        return os.cpu_count()

    now_date = datetime.datetime.now()
    check_date = now_date + datetime.timedelta(days=-1)
    client = IMAPClient(host=config['emails']['setting']['mail_hot'])
    client.login(config['emails']['setting']['username'], config['emails']['setting']['password'])
    client.select_folder('Sent Messages')
    messages = client.search([u'SINCE', check_date])
    sent_message = []

    if messages:
        def process_email(uid):
            response = client.fetch(uid, ['ENVELOPE', 'FLAGS', 'RFC822.SIZE'])
            envelope = response[uid][b'ENVELOPE']
            email_subject, decode = decode_header(envelope.subject.decode())[0]

            if decode:
                try:
                    email_subject = email_subject.decode(decode)
                except LookupError:
                    email_subject = email_subject.decode('gb2312')
            else:
                email_subject = str(email_subject)

            sent_message.append(email_subject)

        if use_multithread:
            with concurrent.futures.ThreadPoolExecutor(max_workers=get_available_threads()) as executor:
                futures = [executor.submit(process_email, uid) for uid in messages]
                concurrent.futures.wait(futures)
        else:
            for uid in messages:
                process_email(uid)

    return param_subject in sent_message


def add_templateID(loci_filt):
    # 先判断属于哪种类型 snp indel 热点, snp和indel只有5列, 热点有8列
    df_filt = loci_filt.drop_duplicates().copy()
    result = (df_filt['ref'].str.len() > 1).any() or (df_filt['alt'].str.len() > 1).any()
    if 'stop' in df_filt.columns and result:
        df_filt['stop'] = df_filt['stop'] + 1
        df_filt['pos'] = df_filt['pos'] - 1
    elif 'stop' in df_filt.columns and not result:
        df_filt['stop'] = df_filt['stop'] + 1
        df_filt['pos'] = df_filt['pos'] - 1
    else:
        if result:
            df_filt['ref'] = df_filt['ref'].astype(str)
            df_filt['alt'] = df_filt['alt'].astype(str)
            df_filt['stop'] = df_filt.apply(lambda row: row['pos'] + max(len(row['ref']), len(row['alt'])) + 1, axis=1)
        else:
            df_filt['stop'] = df_filt['pos'] + 1
        df_filt['pos'] = df_filt['pos'] - 1
    df_filt['TemplateID'] = df_filt['chrom'] + ':' + df_filt['pos'].astype(str) + '-' + df_filt[
        'stop'].astype(
        str)
    df_filt.drop_duplicates('TemplateID', inplace=True)
    return df_filt


def loci_examined_pure(loci_file):
    try:
        if loci_file.split('.')[-1] == 'tsv':
            df_loci = pd.read_csv(loci_file, sep='\t').drop_duplicates()
        elif loci_file.split('.')[-1] == 'xlsx':
            df_loci = pd.read_excel(loci_file).drop_duplicates()
        elif loci_file.split('.')[-1] == 'csv':
            df_loci = pd.read_csv(loci_file).drop_duplicates()
        else:
            logging.error(f'ERROR: loci file: {loci_file}, Unknown file type or does not match primer design file type.')
            sys.exit(1)
    except Exception as e:
        logging.error(f'ERROR: Unable to read WES point selection file {loci_file}, The error log is {e}.')
        sys.exit(1)
    df_loci_snp = df_loci[(df_loci['ref'].str.len() == 1) & (df_loci['alt'].str.len() == 1)].copy()
    df_loci_indel = df_loci[(df_loci['ref'].str.len() > 1) ^ (df_loci['alt'].str.len() > 1)].copy()
    if df_loci.shape[0] < 8:
        # loci < 8 + indel
        message = f'MRD样本ID: {df_loci["sampleSn"].iloc[0]}, 质控结果: SNP位点为: {df_loci_snp.shape[0]} 个, INDEL位点为: {df_loci_indel.shape[0]} 个, SNP+INDEL位点数量为: {df_loci_snp.shape[0] + df_loci_indel.shape[0]}, 当SNP+INDEL数量小于8, 需要审核人员审核退单！'
        logging.error(message)
        sys.exit(0)
    elif 8 <= df_loci.shape[0] < 20:
        # 8 < loci < 20 + 热点
        if not run:
            loci_hots = config['loci_hots']
        else:
            loci_hots = config['local_loci_hots']
        try:
            df_hots = pd.read_excel(loci_hots)
        except Exception as e:
            logging.error(f'ERROR: 获取热点文件失败, 具体原因为: {e}')
            sys.exit(1)
        # snp cancer_type_ID
        loci_cancer_id = list(set(df_loci_snp['cancer_type_ID']))
        if loci_cancer_id == [] or loci_cancer_id[0] == 'unknown':
            logging.error('ERROR: SNP选点的cancer_type_ID为空')
            sys.exit(1)

        def check_id(type_id):
            # 判断cancer_type_ID属于热点文件中CANCER_TYPE_ID哪个cancer tree
            return next(filter(lambda x: type_id.startswith(x), set(df_hots['CANCER_TYPE_ID'])), type_id)

        # snp cancer id
        cancer_res_id = []
        for k in loci_cancer_id:
            cancer_res_id.append(check_id(k))

        # 筛选 + 处理 ——> 热点 + snp
        df_hot = df_hots[df_hots['CANCER_TYPE_ID'].isin(cancer_res_id)].copy().reset_index(drop=True)
        df_hot = df_hot.drop_duplicates(['primer_design_chrom', 'primer_design_start', 'primer_design_end'])
        df_hot_filt = df_hot[
            ['primer_design_chrom', 'primer_design_start', 'primer_design_end', 'Ref', 'Alt', 'Start_Position',
             'Hugo_Symbol', 'End_Position', 'pGVSp', 'cHGVS']].copy()
        df_hot_filt.rename(
            columns={'primer_design_chrom': 'chrom', 'Ref': 'ref', 'Alt': 'alt', 'primer_design_start': 'pos',
                     'Hugo_Symbol': 'gene', 'primer_design_end': 'stop'},
            inplace=True)
        df_hot_filt['vaf'] = np.NaN
        df_hot_filt['chrom'] = df_hot_filt['chrom'].astype(str).apply(lambda x: 'chr' + x)
        df_hot_filt['hots'] = 1

        df_loci['hots'] = 0
        df_snp_hot = pd.concat([df_loci, df_hot_filt]).reset_index(drop=True)

        df_snp_hot[['sampleSn', 'driver', 'cancer_type', 'cancer_type_ID']] = df_snp_hot[
            ['sampleSn', 'driver', 'cancer_type', 'cancer_type_ID']].fillna(method='ffill')

        df_snp_hot['stop'].fillna(df_snp_hot['pos'], inplace=True)
        df_snp_hot['Start_Position'].fillna(df_snp_hot['pos'], inplace=True)
        df_snp_hot['End_Position'].fillna(df_snp_hot['stop'], inplace=True)
        df_snp_hot[["stop", "Start_Position", "End_Position"]] = df_snp_hot[
            ["stop", "Start_Position", "End_Position"]].astype(int)
        df_snp_hot.reset_index(drop=True, inplace=True)
        return df_snp_hot
    else:
        # snp >= 20
        return df_loci


def loci_examined(loci_file, snp, hot, driver):
    try:
        if loci_file.split('.')[-1] == 'tsv':
            try:
                df_loci = pd.read_csv(loci_file, sep='\t').drop_duplicates()
            except UnicodeDecodeError:
                df_loci = pd.read_csv(loci_file, sep='\t', encoding='gbk').drop_duplicates()
        elif loci_file.split('.')[-1] == 'xlsx':
            df_loci = pd.read_excel(loci_file).drop_duplicates()
        elif loci_file.split('.')[-1] == 'csv':
            df_loci = pd.read_csv(loci_file).drop_duplicates()
        else:
            logging.error(f'ERROR: loci file: {loci_file}, Unknown file type or does not match primer design file type.')
            sys.exit(1)
    except Exception as e:
        logging.error(f'ERROR: Unable to read WES point selection file {loci_file}, The error log is {e}.')
        sys.exit(1)
    df_loci_snp = df_loci[(df_loci['ref'].str.len() == 1) & (df_loci['alt'].str.len() == 1)].copy()
    df_loci_indel = df_loci[(df_loci['ref'].str.len() > 1) ^ (df_loci['alt'].str.len() > 1)].copy()
    if df_loci.shape[0] < 8:
        # loci < 8 + indel 即源文件df_loci_filt
        if snp:
            if driver:
                return df_loci
            else:
                if not run:
                    loci_hots = config['loci_hots']
                else:
                    loci_hots = config['local_loci_hots']
                df_hots = pd.DataFrame()
                try:
                    df_hots = pd.read_excel(loci_hots)
                except Exception as e:
                    logging.error(f'ERROR: 获取热点文件失败, 具体原因为: {e}')
                    sys.exit(1)
                # snp cancer_type_ID
                loci_cancer_id = list(set(df_loci_snp['cancer_type_ID']))
                if loci_cancer_id == [] or loci_cancer_id[0] == 'unknown':
                    if DEBUG:
                        subject = f'【MRD引物设计-测试】{df_loci["sampleSn"].iloc[0]} 样本数据错误'
                        toaddrs = config['emails']['toaddrs']['log_toaddrs']
                    else:
                        subject = f'【MRD引物设计】{df_loci["sampleSn"].iloc[0]} 样本数据错误'
                        toaddrs = config['emails']['toaddrs']['error_toaddrs']
                    if not check_emil_sent(subject):
                        emit(
                            message=f'MRD样本ID：{df_loci["sampleSn"].iloc[0]}\n\n数据错误：SNP选点的cancer_type_ID为空！\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                            subject=subject,
                            toaddrs=toaddrs,
                            cc=config['emails']['toaddrs']['log_cc'],
                        )
                    logging.error('ERROR: SNP选点的cancer_type_ID为空')
                    sys.exit(1)

                def check_id(type_id):
                    # 判断cancer_type_ID属于热点文件中CANCER_TYPE_ID哪个cancer tree
                    return next(filter(lambda x: type_id.startswith(x), set(df_hots['CANCER_TYPE_ID'])), type_id)

                # snp cancer id
                cancer_res_id = []
                for k in loci_cancer_id:
                    cancer_res_id.append(check_id(k))

                # 筛选 + 处理 ——> 热点 + snp
                df_hot = df_hots[df_hots['CANCER_TYPE_ID'].isin(cancer_res_id)].copy().reset_index(drop=True)
                df_hot = df_hot.drop_duplicates(['primer_design_chrom', 'primer_design_start', 'primer_design_end'])
                df_hot_filt = df_hot[
                    ['primer_design_chrom', 'primer_design_start', 'primer_design_end', 'Ref', 'Alt', 'Start_Position',
                     'Hugo_Symbol', 'End_Position', 'pHGVS', 'cHGVS']].copy()
                df_hot_filt.rename(
                    columns={'primer_design_chrom': 'chrom', 'Ref': 'ref', 'Alt': 'alt', 'primer_design_start': 'pos',
                             'Hugo_Symbol': 'gene', 'primer_design_end': 'stop'},
                    inplace=True)
                df_hot_filt['vaf'] = np.NaN
                df_hot_filt['chrom'] = df_hot_filt['chrom'].astype(str).apply(lambda x: 'chr' + x)
                df_hot_filt['hots'] = 1

                df_loci['hots'] = 0
                df_snp_hot = pd.concat([df_loci, df_hot_filt]).reset_index(drop=True)

                df_snp_hot[['sampleSn', 'cancer_type', 'cancer_type_ID']] = df_snp_hot[
                    ['sampleSn', 'cancer_type', 'cancer_type_ID']].fillna(method='ffill')

                df_snp_hot['stop'].fillna(df_snp_hot['pos'], inplace=True)
                df_snp_hot['Start_Position'].fillna(df_snp_hot['pos'], inplace=True)
                df_snp_hot['End_Position'].fillna(df_snp_hot['stop'], inplace=True)
                df_snp_hot[["stop", "Start_Position", "End_Position"]] = df_snp_hot[
                    ["stop", "Start_Position", "End_Position"]].astype(int)
                df_snp_hot.reset_index(drop=True, inplace=True)
                return df_snp_hot
        else:
            # 发邮件审核
            if DEBUG:
                subject = f'【MRD引物设计-测试】{df_loci["sampleSn"].iloc[0]} 引物选点未通过质控'
                toaddrs = config['emails']['toaddrs']['log_toaddrs']
            else:
                subject = f'【MRD引物设计】{df_loci["sampleSn"].iloc[0]} 引物选点未通过质控'
                toaddrs = config['emails']['toaddrs']['qc_toaddrs']
            if not check_emil_sent(subject):
                emit(
                    message=f'MRD样本ID：{df_loci["sampleSn"].iloc[0]}\n\n质控结果：\nSNP位点为: {df_loci_snp.shape[0]} 个\nINDEL位点为: {df_loci_indel.shape[0]} 个\nSNP+INDEL位点数量为: {df_loci_snp.shape[0] + df_loci_indel.shape[0]}\n当SNP+INDEL数量小于8,需要审核人员审核退单！\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                    toaddrs=toaddrs,
                    cc=config['emails']['toaddrs']['log_cc'],
                    subject=subject,
                )
            logging.error(f'ERROR: MRD样本ID: {df_loci["sampleSn"].iloc[0]}, SNP+INDEL数量小于8, 已发邮件至审核人员退单!')
            sys.exit(0)
    elif 8 <= df_loci.shape[0] < 20:
        # 8 < loci < 20 + 热点
        # 不加热点
        if not hot:
            if not run:
                loci_hots = config['loci_hots']
            else:
                loci_hots = config['local_loci_hots']
            df_hots = pd.DataFrame()
            try:
                df_hots = pd.read_excel(loci_hots)
            except Exception as e:
                logging.error(f'ERROR: 获取热点文件失败, 具体原因为: {e}')
                sys.exit(1)
            # snp cancer_type_ID
            loci_cancer_id = list(set(df_loci_snp['cancer_type_ID']))
            if loci_cancer_id == [] or loci_cancer_id[0] == 'unknown':
                if DEBUG:
                    subject = f'【MRD引物设计-测试】{df_loci["sampleSn"].iloc[0]} 样本数据错误'
                    toaddrs = config['emails']['toaddrs']['log_toaddrs']
                else:
                    subject = f'【MRD引物设计】{df_loci["sampleSn"].iloc[0]} 样本数据错误'
                    toaddrs = config['emails']['toaddrs']['error_toaddrs']
                if not check_emil_sent(subject):
                    emit(
                        message=f'MRD样本ID：{df_loci["sampleSn"].iloc[0]}\n\n数据错误：SNP选点的cancer_type_ID为空！\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                        subject=subject,
                        toaddrs=toaddrs,
                        cc=config['emails']['toaddrs']['log_cc'],
                    )
                logging.error('ERROR: SNP选点的cancer_type_ID为空')
                sys.exit(1)

            def check_id(type_id):
                # 判断cancer_type_ID属于热点文件中CANCER_TYPE_ID哪个cancer tree
                return next(filter(lambda x: type_id.startswith(x), set(df_hots['CANCER_TYPE_ID'])), type_id)

            # snp cancer id
            cancer_res_id = []
            for k in loci_cancer_id:
                cancer_res_id.append(check_id(k))

            # 筛选 + 处理 ——> 热点 + snp
            df_hot = df_hots[df_hots['CANCER_TYPE_ID'].isin(cancer_res_id)].copy().reset_index(drop=True)
            df_hot = df_hot.drop_duplicates(['primer_design_chrom', 'primer_design_start', 'primer_design_end'])
            df_hot_filt = df_hot[
                ['primer_design_chrom', 'primer_design_start', 'primer_design_end', 'Ref', 'Alt', 'Start_Position',
                 'Hugo_Symbol', 'End_Position', 'pHGVS', 'cHGVS']].copy()
            df_hot_filt.rename(
                columns={'primer_design_chrom': 'chrom', 'Ref': 'ref', 'Alt': 'alt', 'primer_design_start': 'pos',
                         'Hugo_Symbol': 'gene', 'primer_design_end': 'stop'},
                inplace=True)
            df_hot_filt['vaf'] = np.NaN
            df_hot_filt['chrom'] = df_hot_filt['chrom'].astype(str).apply(lambda x: 'chr' + x)
            df_hot_filt['hots'] = 1

            df_loci['hots'] = 0
            df_snp_hot = pd.concat([df_loci, df_hot_filt]).reset_index(drop=True)

            df_snp_hot[['sampleSn', 'cancer_type', 'cancer_type_ID']] = df_snp_hot[
                ['sampleSn', 'cancer_type', 'cancer_type_ID']].fillna(method='ffill')

            df_snp_hot['stop'].fillna(df_snp_hot['pos'], inplace=True)
            df_snp_hot['Start_Position'].fillna(df_snp_hot['pos'], inplace=True)
            df_snp_hot['End_Position'].fillna(df_snp_hot['stop'], inplace=True)
            df_snp_hot[["stop", "Start_Position", "End_Position"]] = df_snp_hot[
                ["stop", "Start_Position", "End_Position"]].astype(int)
            df_snp_hot.reset_index(drop=True, inplace=True)
            return df_snp_hot
        else:
            return df_loci
    else:
        # snp >= 20 直接进行引物设计
        return df_loci


def first_check_driver(df_driver, url, single_result, sampleID):
    if len(df_driver) == 0:
        logging.info(f'此样本无driver基因.')
        return None
    elif len(df_driver) == 1:
        logging.info(f'driver基因位点数量为1, 无需进行单独引物设计.')
        return df_driver['TemplateID'].to_list()
    else:
        logging.info(f'driver基因位点数量为{len(df_driver)}, 进行单独引物设计测试排除兼容性.')
        result_string, not_used = select_site(df_driver)
        headers, cookies, token = primkit.fetch_web_data(method='requests')
        post_data = primkit.prepare_post_data(token, result_string)
        down_url = primkit.design_primers(post_data, method='requests', headers=headers, cookies=cookies)
        # down_url = imitate_web(result_string, url)
        save_name = '{}/{}-{}.csv'.format(os.path.join(single_result, sampleID),
                                          sampleID, 'driver')
        down_result(down_url, save_name)

        try:
            df_res = pd.read_csv(save_name, sep=',', header=3).iloc[0:-1]
        except pd.errors.ParserError:
            with open(save_name, 'r') as f:
                lines = f.readlines()
            # Process the lines
            processed_lines = []
            for i, line in enumerate(lines[4:]):  # Start from the 4th line (zero-indexed)
                parts = line.strip().split(',')
                if len(parts) > 20:
                    parts[-2] = parts[-2] + ';' + parts[-1]
                    del parts[-1]
                    if len(parts) != 20:
                        logging.error(f'Line {i + 5} has an unexpected number of fields ({len(parts)} instead of 20).')
                        continue
                processed_lines.append(','.join(parts))
            # Create a DataFrame from the processed lines
            data = '\n'.join(processed_lines)
            df_res = pd.read_csv(io.StringIO(data), sep=',').iloc[0:-1]

        logging.info('单独引物设计结果为:\n{}'.format(df_res))
        return df_res['TemplateID'].to_list()


def convert_driver_to_string(lst):
    result = ''
    for item in lst:
        chrom, coord = item.split(':')
        start, end = coord.split('-')
        result += f'{chrom}\t{start}\t{end}\n'
    return result


def check_sample_date(sample_id):
    """
    检查样本ID中的日期与当前日期的差距。
    如果日期差距超过10天，则打印警告信息。

    参数:
    sample_id (str): 样本ID，例如 'NGS231115-194WX'

    返回:
    str: 样本日期与当前日期的差距信息
    """
    # 使用正则表达式提取日期
    date_match = re.search(r'NGS(\d{2})(\d{2})(\d{2})-', sample_id)
    if date_match:
        year = int(date_match.group(1)) + 2000  # 假设'23'表示2023年
        month = int(date_match.group(2))
        day = int(date_match.group(3))

        # 构造日期对象
        sample_date = datetime.datetime(year, month, day)
        current_date = datetime.datetime.now()

        # 计算日期差距
        date_difference = (current_date - sample_date).days

        # 检查日期差距是否超过10天
        if date_difference > 10:
            msg = f"警告：样本 {sample_id} 的日期与当前日期相差 {date_difference} 天，超过了10天。"
            emit(message=msg, subject=msg, toaddrs=config['emails']['toaddrs']['log_toaddrs'], cc=config['emails']['toaddrs']['log_cc'])
            return msg
        else:
            return f"样本 {sample_id} 的日期与当前日期相差 {date_difference} 天，未超过10天。"
    else:
        return "无法从样本ID中提取日期。"


def process_file(source_file, url, result_path, snp, driver, hot, check_cms, online=False):
    # 先前检查
    if online:
        file_name = urlparse(source_file).path.split('/')[-1]
    else:
        if '\\' in source_file:
            file_name = source_file.split('\\')[-1]
        else:
            file_name = source_file.split('/')[-1]
    sampleID = file_name.split(".")[0]
    if 'WX' in sampleID:
        sampleID = sampleID.split('WX')[0]
    elif 'WE' in sampleID:
        sampleID = sampleID.split('WE')[0]

    check_sample_date(sampleID)

    if not online and not check_cms:
        # 判断样本ID是否属于MRD项目
        sample_local = determine_sample_location(sampleID)
        if sample_local == 'OLD':
            project_item = get_project_type_old(sampleID)
        elif sample_local == 'NEW':
            project_item = get_project_type(sampleID)
        else:
            logging.error(
                f'ERROR: The sample {sampleID} does not exist in the small wide CMS system, nor does it exist in the thousand wing CMS system')
            sys.exit(1)

        if project_item == 'OTHER':
            if DEBUG:
                subject = f'【MRD引物设计-测试】{sampleID} 不属于迈锐达MRD检测提示'
            else:
                subject = f'【MRD引物设计】{sampleID} 不属于迈锐达MRD检测提示'
            if not check_emil_sent(subject):
                emit(
                    message=f'MRD样本ID：{sampleID}\n\n检测提示：无需进行引物设计，该样本不属于迈锐达MRD检测！\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                    subject=subject,
                    toaddrs=config['emails']['toaddrs']['log_toaddrs'])
            logging.info(
                f'The sample ID is {sampleID} does not belong to the MRD detection of Merida, and no primer design is required')
            sys.exit(0)

    #  make primer outcome path
    single_result = os.path.join(result_path, 'primer_outcome/')
    os.makedirs(os.path.join(single_result, sampleID), exist_ok=True)

    # 判断源文件选点数量
    if online:
        df_loci = loci_examined_pure(source_file)
    else:
        try:
            df_loci = loci_examined(source_file, snp, hot, driver)
        except KeyError as e:
            if DEBUG:
                subject = f'【MRD引物设计-测试】{sampleID} 样本数据错误'
                toaddrs = config['emails']['toaddrs']['log_toaddrs']
            else:
                subject = f'【MRD引物设计】{sampleID} 样本数据错误'
                toaddrs = config['emails']['toaddrs']['error_toaddrs']
            emit(
                message=f'MRD样本ID：{sampleID}\n\n数据错误：{str(e)}! \n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                subject=subject,
                toaddrs=toaddrs,
                cc=config['emails']['toaddrs']['log_cc'],
                )
            logging.error(f'ERROR: 样本 {sampleID} 数据错误')
            sys.exit(1)

    # 添加templateID
    df_loci_handle = add_templateID(df_loci)

    # 对driver数量大于1单独处理一次,筛选出兼容的driver备用
    df_driver = None
    driver_list = []
    driver_str = ''
    if not driver:
        df_driver = df_loci_handle[df_loci_handle['driver'] == 1].copy()
        if not df_driver.empty:
            driver_list = first_check_driver(df_driver, url, single_result, sampleID)
        # 将driver转换为bed字符串
        if driver_list is not None:
            driver_str = convert_driver_to_string(driver_list)
        # 去掉driver基因后进行引物设计
        df_no_driver = df_loci_handle[~(df_loci_handle['driver'] == 1)].copy()
    else:
        df_no_driver = df_loci_handle.copy()

    # 除去能设计的driver, 还需要设计的数量
    if len(driver_list) > 0:
        design_num = 20 - len(driver_list)
    else:
        design_num = 20

    # 开始使用primer3设计引物
    num = 0
    df_res = pd.DataFrame()
    not_used = []
    should_exit = False  # 初始化退出循环的标志
    while True:
        num = num + 1
        logging.info('样本-{}-第{}次引物设计'.format(sampleID, num))
        if df_no_driver.shape[0] >= 20:
            if num == 1:
                result_string, not_used = select_site(df_no_driver, num=design_num)
                if driver_str != '':
                    primer_string = driver_str + result_string
                else:
                    primer_string = result_string
            else:
                logging.info('样本-{}-第 {} 次 iGeneTech 引物设计结果为: {} 条, 未满20条引物, 未使用引物为: {} 条.'.format(sampleID,
                                                                                                   num - 1,
                                                                                                   df_res.shape[0],
                                                                                                   len(not_used)))
                # 由于driver和非driver会把diver剔除，所以遵从算法优先选择
                new_driver_list = df_res[df_res['TemplateID'].isin(driver_list)]['TemplateID'].to_list()
                new_driver_str = convert_driver_to_string(new_driver_list)
                new_design_num = design_num + len(driver_list) - len(new_driver_list)
                if len(new_driver_list) > 0:
                    result_string, not_used = select_site(df_no_driver, df_res, not_used, num=new_design_num,
                                                          driver=new_driver_list)
                else:
                    result_string, not_used = select_site(df_no_driver, df_res, not_used, num=new_design_num)
                if new_driver_str != '':
                    primer_string = new_driver_str + result_string
                else:
                    primer_string = result_string
            # 模拟人工MFEPrimer进行引物设计

            # down_url = imitate_web(primer_string, url)
            headers, cookies, token = primkit.fetch_web_data(method='requests')
            post_data = primkit.prepare_post_data(token, primer_string)
            down_url = primkit.design_primers(post_data, method='requests', headers=headers, cookies=cookies)
            logging.info('样本-{}-第 {} 次 iGeneTech url为: {}.'.format(sampleID, num, down_url))
            save_name = '{}/{}-{}.csv'.format(os.path.join(single_result, sampleID),
                                              sampleID, num)
            # 下载引物结果
            down_result(down_url, save_name)

            # 读取下载完成的文件, 进行质控
            csv_file = '{}/{}-{}.csv'.format(os.path.join(single_result, sampleID), sampleID, num)
            df_res = pd.DataFrame()
            try:
                df_res = pd.read_csv(csv_file, sep=',', header=3).iloc[0:-1]
            except pd.errors.ParserError:
                with open(csv_file, 'r') as f:
                    lines = f.readlines()
                # Process the lines
                processed_lines = []
                for i, line in enumerate(lines[4:]):  # Start from the 4th line (zero-indexed)
                    parts = line.strip().split(',')
                    if len(parts) > 20:
                        parts[-2] = parts[-2] + ';' + parts[-1]
                        del parts[-1]
                        if len(parts) != 20:
                            logging.error(f'Line {i + 5} has an unexpected number of fields ({len(parts)} instead of 20).')
                            continue
                    processed_lines.append(','.join(parts))
                # Create a DataFrame from the processed lines
                data = '\n'.join(processed_lines)
                df_res = pd.read_csv(io.StringIO(data), sep=',').iloc[0:-1]
            finally:
                if df_res.empty:
                    logging.error(f'样本 {sampleID} 读取 iGeneTech 引物设计结果 {csv_file} 失败.')
                else:
                    logging.info(f'样本-{sampleID}-第 {num} 次引物设计结果为:\n{df_res}')

            # if 'chr7:55249070-55249072' in df_res['TemplateID'].to_list():
            #     continue
            # else:
            #     regular_str = "chr7\t55249070\t55249072\n"
            #     result_string, not_used = select_site(df_loci_handle[~(df_loci_handle['TemplateID'] == 'chr7:55249070-55249072')], num=len(df_loci_handle) - 1)
            #
            #     for index, key in enumerate(result_string.split('\n')):
            #         print(key)
            #         regular_str1 = regular_str + key + '\n'
            #         down_url = imitate_web(regular_str1, url)
            #         save_name = '{}/{}-{}.csv'.format(os.path.join(single_result, sampleID),
            #                                           sampleID, index + 50)
            #         down_result(down_url, save_name)
            #         df_res = pd.read_csv(save_name, sep=',', header=3).iloc[0:-1]
            #         if 'chr7:55249070-55249072' in df_res['TemplateID'].to_list():
            #             regular_str += key + '\n'
            #             print('chr7:55249070-55249072 in result')
            #         elif len(df_res) == 20:
            #             should_exit = True
            #         else:
            #             continue
            #
            #     if should_exit or not result_string.split('\n'):  # 检查退出标志或result_string已被完全处理
            #         return save_name, df_loci_handle

            if df_res.shape[0] == 20 or not_used == [] or len(not_used) == 0:
                # 发出警告
                if driver:
                    return save_name, df_loci_handle
                else:
                    if len(df_res) < 20 and not df_driver.empty and len(df_no_driver) > 40:
                        if not online:
                            if DEBUG:
                                subject = f'【MRD引物设计-测试】{sampleID} 选点存在driver导致未满20条引物预警'
                                toaddrs = config['emails']['toaddrs']['log_cc']
                            else:
                                subject = f'【MRD引物设计】{sampleID} 选点存在driver导致未满20条引物预警'
                                toaddrs = config['emails']['toaddrs']['qc_toaddrs']
                            if not check_emil_sent(subject):
                                emit(
                                    message=f'MRD样本ID：{sampleID}\n\n预警信息：\n当样本数据中存在driver基因，且保证基因不干扰情况下尽可能保留driver基因\n此样本选点数量为： {len(df_loci_handle)}个，driver基因数量为：{len(df_driver)}个\n当非driver基因大于40个，但引物设计结果低于20个，发出邮件警告，需要审核及相关人员协商解决问题\n\n解决问题：协商后重新设计引物\n\n\nPS: 2023 MRD PRIMER DESIGN. PLEASE DO NOT REPLY TO SYSTEM EMAIL.\n',
                                    toaddrs=toaddrs,
                                    cc=config['emails']['toaddrs']['log_cc'],
                                    subject=subject,
                                )
                            logging.error(f'ERROR: 样本SampleSn为:{sampleID}, driver存在导致未满20条引物预警已发送!')
                            sys.exit(1)
                        else:
                            logging.error(f'ERROR: 样本SampleSn为:{sampleID}, driver存在导致未满20条引物预警。')
                            sys.exit(1)
                    else:
                        return save_name, df_loci_handle
            else:
                continue
        else:
            result_string, not_used = select_site(df_loci_handle)
            if driver_str != '':
                primer_string = driver_str + result_string
            else:
                primer_string = result_string
            save_name = '{}/{}-{}.csv'.format(os.path.join(single_result, sampleID),
                                              sampleID, num)
            down_url = imitate_web(primer_string, url)
            logging.info('样本-{}-第 {} 次引物设计结果url为:\n{}'.format(sampleID, num, down_url))
            down_result(down_url, save_name)
            return save_name, df_loci_handle


def process_file_with_exception_handling(*args):
    try:
        result_file, df_loci = process_file(*args)
        return result_file, df_loci
    except Exception as e:
        logging.error(f'ERROR: {e}')
        return None, None


def on_process_completion(future, order_path, mold, snp, skip_review):
    try:
        result = future.result()
        if result is not None:
            result_file, df_loci = result
            write_order(result_file, df_loci, order_path, mold, snp, skip_review)
        else:
            logging.error("Error processing file")
    except Exception as e:
        logging.error("Error in on_process_completion: {}.".format(str(e)))


def execute_parallel(mold, source_files, result_path, snp, driver, hot, check_cms, skip_review):
    logging.info('Start primer design:')
    # target url
    url = config['mfe_primer']
    # make order path
    order_path = os.path.join(result_path, 'primer_order')
    if result_path and os.path.exists(result_path):
        os.makedirs(order_path, exist_ok=True)
    else:
        logging.error('Error: result_path does not exist, order_path not created successfully.')
    # Judge whether to use multithreading
    thread = 8  # Set the maximum number of processes
    if isinstance(source_files, list) and len(source_files) > 1:
        with concurrent.futures.ThreadPoolExecutor(max_workers=thread) as executor:
            futures = []
            for source_file in source_files:
                future = executor.submit(process_file_with_exception_handling, source_file, url, result_path, snp,
                                         driver, hot, check_cms)
                future.add_done_callback(lambda future: on_process_completion(future, order_path, mold, snp, skip_review))
                futures.append(future)
            concurrent.futures.wait(futures)
    # url(审核系统调用)
    elif is_url(source_files):
        try:
            result_file, df_loci = process_file(source_files, url, result_path, snp, driver, hot, check_cms, online=True)
            write_order(result_file, df_loci, order_path, mold, snp, skip_review, online=True)
        except Exception as e:
            logging.info(f'ERROR: {e}')
            sys.exit(1)
    # single thread
    else:
        try:
            result_file, df_loci = process_file(source_files, url, result_path, snp, driver, hot, check_cms)
            write_order(result_file, df_loci, order_path, mold, snp, skip_review)
            logging.info('Primer design completed!')
        except Exception as e:
            logging.info(f'ERROR: {e}')
            sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description='Automatic primer design.',
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('-m', '--order_template',
                        dest='mold', type=str, choices=['sh', 'hz', 'sg', 'dg'],
                        help='Currently, the order template is only available in sh(上海百力格), hz(湖州河马), sg(上海生工), dg(上海迪赢).',
                        required=True)
    parser.add_argument('-i', '--input_file',
                        dest='single_file', type=str,
                        help='Primer Design of Single Site File.',
                        required=False)
    parser.add_argument('-I', '--Input_folder',
                        dest='input_folder', type=str,
                        help='Primer design for files containing multiple loci.',
                        required=False)
    parser.add_argument('-u', '--url',
                        dest='url', type=str,
                        help='URL for Primer Design file.',
                        required=False)
    parser.add_argument('-o', '--output_path',
                        dest='output_path', type=str,
                        help='Specify the path of single site file or multiple site file primer design result file.',
                        required=False)
    parser.add_argument('-s', '--snp', dest='snp', action='store_true', default=False,
                        help='Whether to design snp primers.')
    parser.add_argument('-d', '--driver', dest='driver', action='store_true', default=False,
                        help='Whether to design driver primers.')
    parser.add_argument('-H', '--hot', dest='hot', action='store_true', default=False,
                        help='Whether to design hot primers.')
    parser.add_argument('-c', '--check_cms', dest='check_cms', action='store_true', default=False,
                        help='Skip CMS sample detection.')
    parser.add_argument('-r', '--skip_review', dest='skip_review', action='store_true', default=False,
                        help='Skip report review.')
    args = parser.parse_args()

    if not vars(args):
        parser.print_help()
        sys.exit(1)

    if not args.mold:
        logging.error('Error: mold is required.')
    elif args.mold not in ['sh', 'hz', 'sg', 'dg']:
        logging.error('Error: mold must be "sh", "hz", "sg", "dg".')
        sys.exit(1)

    # Check if URL, single_file, and input_folder are mutually exclusive
    sources = [args.url, args.single_file, args.input_folder]
    if sources.count(None) < len(sources) - 1:
        logging.error('Error: url, input_file, and input_folder cannot be used together.')
        sys.exit(1)

    # 判断 url、output_path
    if args.url:
        if not is_url(args.url):
            logging.error('Error: The input URL address does not meet the standard')
        else:
            if args.output_path and not os.path.exists(args.output_path):
                os.makedirs(args.output_path, exist_ok=True)
            elif not args.output_path:
                args.output_path = '/home/ngs/PrimerDesign/primer_design/'
                os.makedirs(args.output_path, exist_ok=True)
            if args.output_path:
                execute_parallel(args.mold, args.url, args.output_path, args.snp, args.driver, args.hot, args.check_cms, args.skip_review)

    # 判断 single_file、output_path
    if args.single_file:
        if not os.path.exists(args.single_file):
            logging.error('Error: input_file does not exist.')
        elif not os.path.isfile(args.single_file):
            logging.error('Error: input_file must be a file, not a directory.')
        else:
            if args.output_path and not os.path.exists(args.output_path):
                os.makedirs(args.output_path, exist_ok=True)
            elif not args.output_path:
                args.output_path = os.path.join(os.path.dirname(args.single_file), '../primer_design/')
                os.makedirs(args.output_path, exist_ok=True)
            if args.output_path:
                execute_parallel(args.mold, args.single_file, args.output_path, args.snp, args.driver, args.hot, args.check_cms, args.skip_review)

    # 判断 input_folder、output_path
    if args.input_folder:
        if not os.path.exists(args.input_folder):
            logging.error('Error: input_folder does not exist.')
        elif not os.path.isdir(args.input_folder):
            logging.error('Error: input_folder must be a directory, not a file.')
        else:
            if args.output_path and not os.path.exists(args.output_path):
                os.makedirs(args.output_path, exist_ok=True)
            elif not args.output_path:
                args.output_path = os.path.join(args.input_folder, '../primer_design')
                os.makedirs(args.output_path, exist_ok=True)
            if args.output_path:
                source_files = [os.path.join(args.input_folder, k) for k in os.listdir(args.input_folder) if
                                k.endswith('tsv')]
                if not source_files:
                    logging.error('Error: input_folder does not contain any tsv files.')
                else:
                    execute_parallel(args.mold, source_files, args.output_path, args.snp, args.driver,
                                     args.hot, args.check_cms, args.skip_review)


if __name__ == '__main__':
    main()
