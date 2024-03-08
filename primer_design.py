#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Project : TopGen
@Time    : 2023/12/28 16:03
@Author  : lbfeng
@File    : primer_design.py
"""
import os
import re
import sys
import json
import time
import pandas as pd
import numpy as np
import primkit as pt
import requests
import logging
import argparse
import yaml
import datetime
import openpyxl
from urllib.parse import urlparse
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 设置日志
logger = logging.getLogger(__name__)

# 读取配置文件
with open('config.yaml', 'r', encoding='utf-8') as f:
    config = yaml.load(f, Loader=yaml.FullLoader)


def emit(subject, message, attachments=None, to_addrs=None, cc_addrs=None, bcc_addrs=None):
    """
    Sends an email with the given parameters.

    :param subject: Subject of the email.
    :param message: Body of the email.
    :param attachments: List of file paths to attach to the email. Default is None.
    :param to_addrs: List of email addresses to send the email to. Default is taken from config.
    :param cc_addrs: List of email addresses for CC. Default is None.
    :param bcc_addrs: List of email addresses for BCC. Default is None.
    """
    email_manager = pt.EmailManager(config['emails']['login'], use_yagmail=True)

    default_to_addrs = config['emails']['setup']['log_toaddrs'] if DEBUG else config['emails']['setup']['qc_toaddrs']
    to_addrs = to_addrs if to_addrs is not None else default_to_addrs

    subject_prefix = '【MRD引物设计-测试】' if DEBUG else '【MRD引物设计】'
    subject = subject_prefix + subject

    email_manager.send_email(to_addrs=to_addrs, subject=subject, message=message, cc_addrs=cc_addrs,
                             bcc_addrs=bcc_addrs, attachments=attachments)


def get_sample_id(file_path):
    """
    Extracts the sample ID from a given file path which can be a URL or a file path.

    :param file_path: The source file URL or file path.
    :return: The extracted sample ID.
    """
    parsed_url = urlparse(file_path)
    file_name = os.path.basename(parsed_url.path) if parsed_url.scheme and parsed_url.netloc else os.path.basename(
        file_path)
    sampleSn = file_name.split('.')[0]
    sampleID = sampleSn[:-2]
    return sampleID


def check_sample_date(sample_id, send_email=True, email_interval=None, exit_threshold=None):
    """
    Checks the difference between the date in the sample ID and the current date.
    If the date difference exceeds the specified number of days, it sends a warning message or exits the program.

    :param sample_id: The sample ID, e.g., 'NGS231115-194WX'.
    :param send_email: Flag indicating whether to send an email.
    :param email_interval: The interval in days at which to send emails.
    :param exit_threshold: If set and the date difference exceeds this value, the program will exit.
    :return: None
    """
    # Extract the date using a regular expression
    date_match = re.search(r'NGS(\d{2})(\d{2})(\d{2})-', sample_id)
    if date_match:
        year = int(date_match.group(1)) + 2000  # Assuming '23' refers to the year 2023
        month = int(date_match.group(2))
        day = int(date_match.group(3))

        # Construct the date object
        sample_date = datetime.datetime(year, month, day)
        current_date = datetime.datetime.now()

        # Calculate the date difference
        date_difference = (current_date - sample_date).days

        # Check if the date difference exceeds the specified interval
        if exit_threshold and date_difference > exit_threshold:
            msg = f"Error: The sample {sample_id} date differs from the current date by {date_difference} days, exceeding the threshold of {exit_threshold} days. The program will exit."
            logger.critical(msg)
            if send_email:
                subject = f"样本日期检查严重警告 - {sample_id}"
                message = f"严重警告：样本ID {sample_id} 日期检查\n样本日期: {sample_date}\n当前日期: {current_date}\n日期差距: {date_difference} 天\n警告：样本日期与当前日期的差距已超过阈值 {exit_threshold} 天。\n程序将自动退出以防止进一步的数据处理。\n请立即检查相关数据并采取适当措施。"
                emit(subject, message)
            sys.exit(1)

        if date_difference > email_interval:
            msg = f"Warning: The sample {sample_id} date differs from the current date by {date_difference} days, exceeding the interval of {email_interval} days."
            logger.warning(msg)
            if send_email:
                subject = f"样本日期检查警告 - {sample_id}"
                message = f"警告：样本ID {sample_id} 日期检查\n样本日期: {sample_date}\n当前日期: {current_date}\n日期差距: {date_difference} 天\n提示：样本日期与当前日期的差距已超过设定的 {email_interval} 天。\n请检查相关数据以确保样本的有效性和时效性。"
                emit(subject, message)
        else:
            logger.info(
                f"The sample {sample_id} date differs from the current date by {date_difference} days, which does not exceed the interval of {email_interval} days.")
    else:
        logger.warning("Unable to extract the date from the sample ID.")


def get_cms_accessToken():
    headers = config['header']
    postUrl = config['CMS_URL']['accessToken']['post_url']
    postData = config['CMS_URL']['accessToken']['post_data']
    try:
        response = requests.post(postUrl, params=postData, headers=headers)
        if response.status_code == 200:
            accessToken = json.loads(response.text)['data']['accessToken']
            logger.info(f'AccessToken obtained successfully, accessToken: {accessToken}')
            return accessToken
        else:
            logger.error('ERROR: response.status_code is not equal to 200.')
            sys.exit(1)
    except Exception as e:
        logger.error(f'ERROR: Error getting sample audit status of cms system. The specific reason is {e}')
        sys.exit(1)


def get_project_name(sampleSn):
    """
    Retrieves the project name for a given sample ID based on its location.

    :param sampleSn: The sample ID to check.
    :return: The project name associated with the sample ID.
    """
    # Code for the old system
    accessToken = get_cms_accessToken()
    sampleInfo_url = config['CMS_URL']['sampleInfo']['get_url']
    payload = {'accessToken': accessToken, "search[sampleSn][value]": sampleSn, "search[sampleSn][query]": "eq"}

    try:
        result = requests.get(sampleInfo_url, params=payload)
        dicts = json.loads(result.text)
        if len(dicts["data"]) > 0:
            project_id = dicts["data"][0]["itemName"]
            logger.info(f'Successfully obtained project ID, project ID: {project_id}')
            return project_id
        else:
            logger.error('ERROR: Sample ID does not exist in cms system!')
            sys.exit(1)
    except Exception as e:
        logger.error(f'ERROR: An error occurred while getting the project name for the sample. Reason: {e}')
        sys.exit(1)


def get_project_type(project_name):
    """
    Determines if the project type is MRD based on the project name and sample location.

    :param project_name: The name of the project.
    :return: True if the project type is 'MRD', False otherwise.
    """
    try:
        project_id_match = re.search(r"^(.{7})", project_name)
        if project_id_match:
            project_id = project_id_match.group(1)
            MRD_detection = config['MRD_ID']
            is_MRD = project_id in MRD_detection
            if is_MRD:
                logger.info(f"Project {project_name} is of type 'MRD'.")
            else:
                logger.warning(f"Project {project_name} is not of type 'MRD'.")
            return is_MRD
        else:
            logger.warning(f"Project name {project_name} does not match the expected pattern.")
            return False
    except Exception as e:
        logger.error(f"Error determining project type for {project_name}: {e}")
        return False


def get_sample_status(sampleSn):
    """
    Retrieves the status of a sample based on the specified system type.

    :param sampleSn: The sample number to check the status for.
    :return: The status of the sample.
    """
    access_token = get_cms_accessToken()
    sample_info_url = config['CMS_URL']['sampleInfo']['get_url']
    payload = {'accessToken': access_token, "search[sampleSn][value]": sampleSn, "search[sampleSn][query]": "eq"}
    try:
        result = requests.get(sample_info_url, params=payload)
        dicts = json.loads(result.text)
        if len(dicts["data"]) > 0:
            sample_status = dicts["data"][0]["sampleStatusShow"]
            logging.info(f'Successfully obtained sampleStatus, sampleStatus: {sample_status}')
            return sample_status
        else:
            logger.error('ERROR: Sample ID does not exist in cms system!')
            sys.exit(1)
    except Exception as e:
        logger.error(
            f'ERROR: An error occurred while getting the item ID of the sample for the cms system. The specific reason is {e}')
        sys.exit(1)


def get_audit_status(sampleSn):
    """
    Retrieves the audit status for a given sample number.

    :param sampleSn: The sample number to check the status for.
    :return: A tuple containing the audit status code and its description.
    """
    status_dict = config['review_status']
    audit_status = get_sample_status(sampleSn)

    # For the OLD system, find the abbreviation from the status description
    status_abbr = next((abbr for abbr, desc in status_dict.items() if desc == audit_status), 'Unknown')
    status_desc = audit_status

    return status_abbr, status_desc


def handle_mrd_sample(sampleID, send_email=True):
    """
    Handles the MRD sample ID by determining its project type and sending email notifications if necessary.

    :param sampleID: The MRD sample ID to be processed.
    :param send_email: Flag indicating whether to send an email.
    """
    project_name = get_project_name(sampleID)
    project_type = get_project_type(project_name)

    if not project_type:
        logger.error(
            f'The sample ID {sampleID} does not belong to the MRD detection, and no primer design is required.')
        if send_email:
            subject = f"样本项目检查警告 - {sampleID}"
            message = f"警告：样本ID {sampleID} 项目检查\n提示：样本不属于迈锐达检测，不进行引物设计。\n请检查样本项目配置。"
            emit(subject, message)
        sys.exit(0)


def read_loci_file(file_path):
    """
    Reads a file into a pandas DataFrame. Supports TSV, XLSX, CSV, and TXT formats with automatic delimiter detection for text files.

    :param file_path: Path to the file to be read.
    :return: A pandas DataFrame containing the file's content.
    """
    try:
        file_extension = file_path.split('.')[-1].lower()

        if file_extension in ['xlsx']:
            df = pd.read_excel(file_path).drop_duplicates()
        elif file_extension in ['csv', 'tsv', 'txt']:
            # For CSV, TSV, and TXT, try to automatically detect delimiter
            try:
                df = pd.read_csv(file_path, sep=None, engine='python').drop_duplicates()
            except UnicodeDecodeError:
                df = pd.read_csv(file_path, sep=None, engine='python', encoding='gbk').drop_duplicates()
        else:
            logger.error(f'ERROR: Unknown file type or does not match expected file types for file: {file_path}.')
            sys.exit(1)

        return df

    except Exception as e:
        logger.error(f'ERROR: Unable to read file {file_path}, The error log is {e}.')
        sys.exit(1)


def read_hots_file():
    """
    Reads a hotspots file specified in the configuration into a pandas DataFrame.

    :return: A pandas DataFrame containing the hotspots data.
    """
    loci_hots = config['loci_hots']
    try:
        df_hots = pd.read_excel(loci_hots)
        cancer_ids = df_hots['CANCER_TYPE_ID'].unique().tolist()
        return df_hots, cancer_ids
    except Exception as e:
        logger.error(f'ERROR: Failed to retrieve the hotspots file, the specific reason is: {e}')
        sys.exit(1)


def validate_cancer_type(df_snp, hots_cancer_ids, cancer_id=None):
    """
    Validates if the DataFrame has 'cancer_type_ID' column or uses the provided cancer_id.

    :param df_snp: The DataFrame containing SNP loci data.
    :param hots_cancer_ids: cancer type ID of hotspot file
    :param cancer_id: Default cancer type ID if not present in the DataFrame.
    :return: None
    """

    def check_id(type_id):
        # 判断cancer_type_ID属于热点文件中CANCER_TYPE_ID哪个cancer tree
        return next(filter(lambda x: type_id.startswith(x), hots_cancer_ids), type_id)

    if 'cancer_type_ID' in df_snp.columns:
        loci_cancer_id = list(set(df_snp['cancer_type_ID']))
        cancer_res_id = [check_id(k) for k in loci_cancer_id]
        if not loci_cancer_id or loci_cancer_id[0] == 'unknown':
            sampleSn = df_snp["sampleSn"].iloc[0] if "sampleSn" in df_snp.columns else "UnknownSample"
            subject = f'样本 cancer_type_ID 检查警告 - {sampleSn}'
            message = f'警告：样本ID {sampleSn} 检查 cancer_type_ID 为 unknown。\n提示：样本 cancer_type_ID 为空（unknown），无法为其样本增加热点引物 。\n请检查相关数据以确保样本的准确性。'
            emit(subject, message)
    elif cancer_id:
        loci_cancer_id = [cancer_id]
        cancer_res_id = [check_id(k) for k in loci_cancer_id]
    else:
        logger.error('ERROR: "cancer_type_ID" is not in the DataFrame and no default cancer_id was provided.')
        sys.exit(1)

    return cancer_res_id


def process_hotspots(df_hots, df_loci, cancer_res_id):
    """
    Processes hotspots and loci DataFrames and merges them based on cancer research IDs.

    :param df_hots: DataFrame containing hotspots data.
    :param df_loci: DataFrame containing loci data.
    :param cancer_res_id: List of cancer research IDs.
    :return: A DataFrame with the combined and processed data.
    """

    # Filter hotspots based on cancer research IDs and remove duplicates
    df_hot = df_hots[df_hots['CANCER_TYPE_ID'].isin(cancer_res_id)].drop_duplicates(
        ['primer_design_chrom', 'primer_design_start', 'primer_design_end'])

    # Prepare df_hot for merging
    df_hot = df_hot.rename(columns={
        'primer_design_chrom': 'chrom',
        'primer_design_start': 'pos',
        'primer_design_end': 'stop',
        'Ref': 'ref',
        'Alt': 'alt',
        'Hugo_Symbol': 'gene'
    })
    df_hot = df_hot.assign(vaf=np.NaN, hots=1)
    df_hot['chrom'] = df_hot['chrom'].astype(str).apply(lambda x: 'chr' + x)

    # Select relevant columns for merging
    columns = ['chrom', 'pos', 'stop', 'ref', 'alt', 'Start_Position', 'End_Position', 'pHGVS', 'cHGVS', 'gene', 'vaf',
               'hots']
    df_hot = df_hot[columns]

    # Mark non-hotspots in df_loci
    df_loci = df_loci.assign(hots=0)

    # Combine the hotspots and loci data
    df_combined = pd.concat([df_loci, df_hot], ignore_index=True)

    # Fill missing values and adjust data types
    df_combined['stop'].fillna(df_combined['pos'], inplace=True)
    df_combined['Start_Position'].fillna(df_combined['pos'], inplace=True)
    df_combined['End_Position'].fillna(df_combined['stop'], inplace=True)
    df_combined[['stop', 'Start_Position', 'End_Position']] = df_combined[
        ['stop', 'Start_Position', 'End_Position']].astype(int)

    # Forward fill missing values for certain columns
    fill_columns = ['sampleSn', 'cancer_type', 'cancer_type_ID']
    df_combined[fill_columns] = df_combined[fill_columns].fillna(method='ffill')

    return df_combined


def loci_examined(df_loci, skip_snp_design, skip_hot_design, skip_driver_design, cancer_id=None, send_email=True):
    """
    Examines loci in a given DataFrame and performs various checks and processes based on the parameters provided.

    :param df_loci: DataFrame containing loci information.
    :param skip_snp_design: Boolean flag to skip SNP design.
    :param skip_hot_design: Boolean flag to skip hot design.
    :param skip_driver_design: Boolean flag to skip driver design.
    :param cancer_id: Optional cancer ID for further analysis.
    :param send_email: Flag indicating whether to send an email.

    :return: Processed DataFrame based on the given parameters and conditions.
    """

    # Sample ID
    sampleSn = df_loci['sampleSn'].iloc[0] if 'sampleSn' in df_loci.columns else None

    # Count SNP and INDEL and total loci
    loci_count = df_loci.shape[0]
    df_loci_snp = df_loci[(df_loci['ref'].str.len() == 1) & (df_loci['alt'].str.len() == 1)].copy()
    snp_count = df_loci_snp.shape[0]
    df_loci_indel = df_loci[(df_loci['ref'].str.len() > 1) ^ (df_loci['alt'].str.len() > 1)].copy()
    indel_count = df_loci_indel.shape[0]

    # 仅当 loci_count 小于 20 时读取热点信息
    def process_hotspots_logic():
        df_hots, cancer_ids = read_hots_file()
        if cancer_id and cancer_id not in cancer_ids:
            logger.error(f'ERROR: The cancer_id "{cancer_id}" is not present in the HOTS file.')
            sys.exit(1)
        cancer_res_id = validate_cancer_type(df_loci, cancer_ids, cancer_id)
        return process_hotspots(df_hots, df_loci, cancer_res_id)

    # Decision-making
    if loci_count < 8:
        if skip_snp_design:
            if not (skip_driver_design or skip_hot_design):
                return process_hotspots_logic()
            return df_loci
        else:
            logging.error(f'样本ID: {sampleSn}, SNP + INDEL数量小于8, 已发邮件至审核人员处理！')
            if send_email:
                subject = f'样本位点数量检查警告 - {sampleSn}'
                message = f'警告：样本ID {sampleSn} 位点数量不足。\n质控结果：SNP位点为: {snp_count} 个，INDEL位点为: {indel_count} 个，SNP+INDEL位点数量为: {snp_count + indel_count} 。\n提示：当 SNP + INDEL 数量小于8，需要审核人员审核处理！\n'
                emit(subject, message)
            sys.exit(0)
    elif 8 <= loci_count < 20:
        if not skip_hot_design:
            return process_hotspots_logic()
    return df_loci


def add_templateID(df_loci):
    """
    Processes a DataFrame of genetic loci to add a TemplateID based on chromosome position and type (SNP, INDEL, hotspot).

    :param df_loci: DataFrame containing loci information.
    :return: DataFrame with added TemplateID and adjusted positions.
    """
    # Drop duplicates and make a copy
    df_dup = df_loci.drop_duplicates().copy()

    # Check if there are INDELs
    has_indel = (df_dup['ref'].str.len() > 1).any() or (df_dup['alt'].str.len() > 1).any()

    # Adjust 'stop' and 'pos' based on the presence of 'stop' column and INDELs
    if 'stop' in df_dup.columns:
        df_dup['stop'] += 1
        df_dup['pos'] -= 1
    else:
        if has_indel:
            df_dup['stop'] = df_dup.apply(lambda row: row['pos'] + max(len(row['ref']), len(row['alt'])), axis=1)
        else:
            df_dup['stop'] = df_dup['pos'] + 1
        df_dup['pos'] -= 1

    # Create TemplateID
    df_dup['TemplateID'] = df_dup['chrom'] + ':' + df_dup['pos'].astype(str) + '-' + df_dup['stop'].astype(str)

    # Drop duplicates based on TemplateID
    df_dup.drop_duplicates('TemplateID', inplace=True)

    return df_dup


def select_site(df_source, df_res=None, not_used=None, num=20, driver=None):
    """
    Selects sites from a source DataFrame and handles unused and driver sites.

    :param df_source: DataFrame containing source data.
    :param df_res: DataFrame containing results data.
    :param not_used: List of TemplateIDs not used.
    :param num: Number of sites to select.
    :param driver: List of driver TemplateIDs.
    :return: A string of selected site information and a list of not used TemplateIDs.
    """

    def convert_row_to_string(row):
        row_string = "\t".join(str(x) for x in row)
        return row_string + "\n"

    # Function to handle selection and conversion of data
    def handle_selection(df, number):
        selected = df.head(number).copy()
        used_ids = selected['TemplateID'].to_list()
        unused_ids = df[~df['TemplateID'].isin(used_ids)]['TemplateID'].to_list()
        result_str = "".join(selected[['chrom', 'pos', 'stop']].apply(convert_row_to_string, axis=1))[:-1]
        return result_str, unused_ids

    if df_res is None and not_used is None:
        return handle_selection(df_source, num)
    else:
        df_filt = df_res if driver is None else df_res[~df_res['TemplateID'].isin(driver)]
        # Combine successfully used and not used
        all_used = df_filt['TemplateID'].to_list() + (not_used or [])
        df_filtered = df_source[df_source['TemplateID'].isin(all_used)].drop_duplicates('TemplateID', keep='first')

        return handle_selection(df_filtered, num)


def design_primers_core(url, outcome_dir, sampleID, result_string, file_suffix='driver'):
    """
    Core function for primer design. It selects sites, fetches web data, prepares and posts data for primer design,
    downloads the results, and reads the resulting data.

    :param url: URL for the web service for primer design.
    :param outcome_dir: Directory to save the outcome files.
    :param sampleID: Sample ID for the primer design.
    :param result_string: Format string for primer design.
    :param file_suffix: Suffix for the file name.
    :return: DataFrame containing the results of primer design.
    """

    # Ensure the sampleID directory exists
    sample_dir = os.path.join(outcome_dir, sampleID)
    if not os.path.exists(sample_dir):
        os.makedirs(sample_dir)

    PRIMER_PARAMS = config['PRIMER_PARAMS']

    # Select sites and prepare data for posting
    headers, cookies, token = pt.fetch_web_data(url=url, method='requests')
    post_data = pt.prepare_post_data(token, result_string, custom_params=PRIMER_PARAMS)

    # Design primers and download the results
    down_url = pt.design_primers(post_data, method='requests', headers=headers, cookies=cookies)
    save_path = os.path.join(sample_dir, f'{sampleID}-{file_suffix}.csv')
    pt.download(down_url, save_path)

    # Read and log the result
    file_reader = pt.FileReader()
    df_res = file_reader.read_csv(save_path)

    # Add a column to distinguish file_suffix results and sampleID
    df_res.insert(0, 'sampleID', len(df_res) * [sampleID])
    df_res['Suffix'] = file_suffix

    logger.info(f'第 {file_suffix} 次引物设计结果为:\n{df_res}')

    return df_res, save_path


def save_to_database(df_res, table_name):
    """
    Saves the given DataFrame to a database table specified in the configuration or the provided table name.

    :param df_res: DataFrame to be saved in the database.
    :param table_name: The name of the table where the DataFrame will be saved.
    """
    db_handler.create_df_table(table_name, df_res)
    db_handler.insert_df(table_name, df_res)


def first_check_driver(df_driver, url, outcome_dir, sampleID):
    """
    Checks the number of driver genes in the given DataFrame and performs actions accordingly.

    :param df_driver: DataFrame containing driver gene information.
    :param url: URL for the web service for primer design.
    :param outcome_dir: Directory to save the outcome files.
    :param sampleID: Sample ID for the driver check.
    :return: List of TemplateIDs or None.
    """
    driver_count = df_driver.shape[0]

    if driver_count == 0:
        logger.info(f'提示：样本ID {sampleID} 选点文件中无driver基因。')
        return None

    if driver_count == 1:
        logger.info(f'提示：样本ID {sampleID} 选点文件中driver基因数量为1，无需进行单独引物设计。')
        return df_driver['TemplateID'].to_list()

    logger.info(f'提示：样本ID {sampleID} 选点文件中driver基因数量为{driver_count}，进行单独引物设计测试排除兼容性。')

    result_string, not_used = select_site(df_driver)
    df_res, save_path = design_primers_core(url, outcome_dir, sampleID, result_string, file_suffix='driver')

    # Save the DataFrame to a table in the database.
    save_to_database(df_res, 'mfe_primers')

    return df_res['TemplateID'].to_list()


def process_driver(df_loci, url, outcome_dir, sampleID, skip_driver_design):
    """
    Process the provided DataFrame to filter out driver genes, calculate the number of designs needed,
    and generate a list of driver genes.

    :param df_loci: DataFrame containing loci information.
    :param url: URL used in first_check_driver function.
    :param outcome_dir: Parameter used in first_check_driver function.
    :param sampleID: Sample ID used in first_check_driver function.
    :param skip_driver_design: Boolean flag to indicate if driver is to be considered.
    :return: Tuple containing DataFrame without driver genes, number of designs needed, and list of driver genes.
    """

    def convert_driver_to_string(lst):
        return ''.join(
            [f"{item.split(':')[0]}\t{item.split(':')[1].split('-')[0]}\t{item.split(':')[1].split('-')[1]}\n" for item
             in lst])

    if skip_driver_design or df_loci['driver'].sum() == 0:
        return df_loci, 20, [], ''

    driver_list = first_check_driver(df_loci[df_loci['driver'] == 1], url, outcome_dir, sampleID)
    driver_str = convert_driver_to_string(driver_list)

    df_no_driver = df_loci[~(df_loci['driver'] == 1)]
    design_num = max(20 - len(driver_list), 0)

    return df_no_driver, design_num, driver_list, driver_str


def update_primer_design(df_res, driver_list, design_num):
    """
    Updates the number of designs needed and driver list based on the current results.

    :param df_res: DataFrame with the current primer design results.
    :param driver_list: List of driver gene TemplateIDs.
    :param design_num: Initial number of designs needed.
    :return: Tuple of updated number of designs needed and updated driver list.
    """
    current_drivers = df_res[df_res['TemplateID'].isin(driver_list)]['TemplateID'].to_list()
    updated_design_num = design_num + len(driver_list) - len(current_drivers)

    return updated_design_num, current_drivers


def select_site_logic(df_no_driver, df_res, not_used, design_num, driver_list, driver_str, num):
    """
    Logic for selecting sites for primer design.
    """
    if num == 1:
        result_string, not_used = select_site(df_no_driver, num=design_num)
        primer_string = driver_str + result_string if driver_str else result_string
    else:
        new_design_num, new_driver_list = update_primer_design(df_res, driver_list, design_num)
        result_string, not_used = select_site(df_no_driver, df_res, not_used, num=new_design_num,
                                              driver=new_driver_list)
        primer_string = driver_str + result_string if driver_str else result_string

    return primer_string, not_used


def should_exit_loop(df_res, not_used):
    """
    Determine if the loop should exit based on the results.

    :param df_res: DataFrame with the current primer design results.
    :param not_used: List of TemplateIDs not used in the current primer design.
    :return: Boolean indicating whether to exit the loop.
    """
    if df_res.shape[0] == 20 or not not_used or len(not_used) == 0:
        return True
    return False


def perform_primer_design(df_no_driver, sampleID, url, outcome_dir, design_num, driver_list, driver_str):
    """
    Perform primer design based on the given data.
    """
    num = 0
    df_res = pd.DataFrame()
    not_used = []

    while True:
        num += 1
        logger.info(f'样本 - {sampleID} 第 {num} 次引物设计')

        # Select sites for primer design
        result_string, not_used = select_site_logic(df_no_driver, df_res, not_used, design_num, driver_list, driver_str,
                                                    num)

        # Design primers and process results
        df_res, save_path = design_primers_core(url, outcome_dir, sampleID, result_string, file_suffix=str(num))

        # Save the DataFrame to a table in the database.
        save_to_database(df_res, 'mfe_primers')

        if should_exit_loop(df_res, not_used):
            break

    return df_res


def process_primer_results(df_res, df_design, sampleID, skip_snp_design, send_email=True):
    """
    Processes primer results for quality control and prepares the final data frame.

    :param df_res: DataFrame containing primer results.
    :param df_design: DataFrame containing source data.
    :param sampleID: Sample ID.
    :param skip_snp_design: Boolean flag to skip SNP design.
    :param send_email: Flag indicating whether to send an email.
    :return: Processed DataFrame df_sample.
    """

    def send_quality_control_email(reason, sample_count):
        subject = f'样本引物结果检查警告 - {sampleID}'
        message = f"警告：样本ID {sampleID} 引物结果检查\n质控结果：{reason}数量为 {sample_count}\n提示：当引物结果数量小于12个或是自身位点小于8个时，不满足质控要求，需要审核人员审核处理。\n程序将自动退出以防止进一步的数据处理。\n请立即检查相关数据并采取适当措施。"
        if send_email:
            emit(subject, message)
        logging.error(f'样本ID：{sampleID}, 引物结果未通过质控，请立即检查相关数据并采取适当措施！')
        sys.exit(1)

    if not skip_snp_design:
        if df_res.shape[0] < 12:
            send_quality_control_email('引物结果', df_res.shape[0])
        elif 'hots' in df_design.columns:
            df_hots = pd.merge(df_res[['TemplateID']], df_design[['TemplateID', 'hots']],
                               on='TemplateID').drop_duplicates('TemplateID', keep='first')
            if df_hots[df_hots['hots'] == 0].shape[0] < 8:
                send_quality_control_email('自身位点', df_hots[df_hots['hots'] == 0].shape[0])

    if not skip_snp_design:
        if df_res.shape[0] < 12:
            send_quality_control_email('引物结果', df_res.shape[0])
        elif 'hots' in df_design.columns:
            df_hots = pd.merge(df_res[['TemplateID']], df_design[['TemplateID', 'hots']],
                               on='TemplateID').drop_duplicates('TemplateID', keep='first')
            if df_hots[df_hots['hots'] == 0].shape[0] < 8:
                send_quality_control_email('自身位点', df_hots[df_hots['hots'] == 0].shape[0])

    if 'hots' in df_design.columns and 'Start_Position' in df_design.columns and 'pos' in df_design.columns:
        df_design['pos'] = df_design['Start_Position']
    elif 'pos' in df_design.columns and 'Start_Position' not in df_design.columns:
        df_design['pos'] = df_design['pos'] + 1

    # Prevent the hotspot design from having a gap of 3 columns with no hotspot design, which may cause problems when inserting into the database
    for column in ['hots', 'Start_Position', 'End_Position']:
        if column not in df_design.columns:
            df_design[column] = None
    save_to_database(df_design, 'mrd_selection')

    df_sample = pd.merge(df_res, df_design, on='TemplateID').drop_duplicates('TemplateID', keep='first')

    return df_sample


def process_primer_sample(df_sample):
    """
    Processes the primer sample data by rearranging columns and adjusting formats.

    :param df_sample: DataFrame containing primer sample data.
    :return: Processed DataFrame with specified columns in front.
    """
    # Calculate F_id, R_id and primerID
    df_sample['F_id'] = df_sample.apply(
        lambda row: f'P{row["sampleSn"].split("NGS")[-1].split("W")[0]}-{row.name + 1:02d}F', axis=1)
    df_sample['R_id'] = df_sample.apply(
        lambda row: f'P{row["sampleSn"].split("NGS")[-1].split("W")[0]}-{row.name + 1:02d}R', axis=1)
    df_sample['primerID'] = df_sample.apply(
        lambda row: f'P{row["sampleSn"].split("NGS")[-1].split("W")[0]}-{row.name + 1:02d}', axis=1)

    # Convert primer sequences to uppercase
    df_sample['ForwardPrimer(Fp)'] = df_sample['ForwardPrimer(Fp)'].apply(lambda x: x.upper())
    df_sample['ReversePrimer(Rp)'] = df_sample['ReversePrimer(Rp)'].apply(lambda x: x.upper())

    # Add Selected columns and calculate generic sequence sizes
    df_sample['Selected'] = 1
    df_sample['GnlAmpSize (bp)'] = df_sample['AmpSize(bp)'] + 140

    # Add failed_reason columns
    df_sample['failed_reason'] = ''

    # Specified columns
    specified_columns1 = ['sampleSn', 'chrom', 'pos', 'ref', 'alt', 'Selected', 'failed_reason']
    specified_columns2 = ['F_id', 'R_id', 'primerID', 'GnlAmpSize (bp)']

    # Gets columns other than those specified
    other_columns = [col for col in df_sample.columns if col not in specified_columns1 + specified_columns2]

    # Arrange the specified columns first, followed by the other columns
    df_sample = df_sample[specified_columns1 + specified_columns2 + other_columns]

    save_to_database(df_sample, 'primer_combined')

    return df_sample


def process_primer_order(df, mold):
    """
    Process the given DataFrame by combining forward and reverse primers information,
    adding additional columns, and saving the processed DataFrame to a database.

    :param df: DataFrame to be processed. It should contain columns 'sampleSn', 'F_id',
               'ForwardPrimer(Fp)', 'R_id', and 'ReversePrimer(Rp)'.
    :param mold: Name of the ordering company to be added to the 'OrderingCompany' column.
    :return: Processed DataFrame with combined primer information and additional columns.
    """
    f_sequence_prefix = config['f_sequence_prefix']
    r_sequence_prefix = config['r_sequence_prefix']
    TubeCount = config['TubeCount']
    TotalQuantityOD = config['TotalQuantityOD']
    PurificationMethod = config['PurificationMethod']
    Nmoles = config['Nmoles']
    Modification5Prime = config['Modification5Prime']
    Modification3Prime = config['Modification3Prime']
    DualLabelModification = config['DualLabelModification']
    Remarks = config['Remarks']

    # Process forward primers
    df_primers_f = df[['sampleSn', 'F_id', 'ForwardPrimer(Fp)']].rename(
        columns={'F_id': 'PrimerName', 'ForwardPrimer(Fp)': 'Sequence'})
    df_primers_f['Sequence'] = df_primers_f['Sequence'].apply(lambda x: f_sequence_prefix + x)

    # Process reverse primers
    df_primers_r = df[['sampleSn', 'R_id', 'ReversePrimer(Rp)']].rename(
        columns={'R_id': 'PrimerName', 'ReversePrimer(Rp)': 'Sequence'})
    df_primers_r['Sequence'] = df_primers_r['Sequence'].apply(lambda x: r_sequence_prefix + x)

    # Combine processed DataFrames
    df_combined = pd.concat([df_primers_f, df_primers_r]).sort_values('sampleSn').reset_index(drop=True)

    # Add additional columns
    df_combined['BaseCount'] = df_combined['Sequence'].apply(len)
    df_combined['TubeCount'] = TubeCount
    df_combined['TotalQuantityOD'] = TotalQuantityOD
    df_combined['PurificationMethod'] = PurificationMethod
    df_combined['Nmoles'] = Nmoles
    df_combined['Modification5Prime'] = Modification5Prime
    df_combined['Modification3Prime'] = Modification3Prime
    df_combined['DualLabelModification'] = DualLabelModification
    df_combined['Remarks'] = Remarks
    df_combined['OrderingCompany'] = mold
    df_combined['DesignDate'] = pd.to_datetime(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    df_combined.sort_values('PrimerName', inplace=True)
    df_combined.reset_index(drop=True, inplace=True)

    # Save to database
    save_to_database(df_combined, 'primer_order')

    return df_combined


def write_sh_order(df_sample, dataframe, order_path, sampleID):
    os.makedirs(order_path, exist_ok=True)
    try:
        order_template = config['order_template']['sh']
        wb = openpyxl.load_workbook(order_template)
    except Exception as e:
        logging.error(f'ERROR: {e}')
        sys.exit(1)

    ws_dna = wb['DNA合成订购表']

    def write_to_sheet(data, column, start_row, sequence_prefix=None):
        for i, value in enumerate(data):
            if value.islower():
                value = value.upper()
            if sequence_prefix:
                value = sequence_prefix + value
            ws_dna.cell(start_row + i, column).value = value
            ws_dna.cell(start_row + i, 5).value = 'PAGE'
            ws_dna.cell(start_row + i, 8).value = 1
            ws_dna.cell(start_row + i, 10).value = 2
            ws_dna.cell(start_row + i, 11).value = 'BLG白色标签'
            ws_dna.cell(start_row + i, 12).value = '1.5ml离心管'
            ws_dna.cell(start_row + i, 13).value = 50
            ws_dna.cell(start_row + i, 14).value = 'H+M'
            ws_dna.cell(start_row + i, 15).value = '1*TE 稀释'

    # Write the 'F_id' data to the sheet
    write_to_sheet(df_sample['F_id'], 3, 20)

    # Leave a blank row after writing the 'F_id' data
    ws_dna.cell(20 + len(df_sample['F_id']), 3).value = None

    # Write the 'R_id' data to the sheet
    write_to_sheet(df_sample['R_id'], 3, 21 + len(df_sample['F_id']))

    # Write the 'ForwardPrimer(Fp)' data to the sheet
    write_to_sheet(df_sample['ForwardPrimer(Fp)'], 4, 20, "GTTCAGAGTTCTACAGTCCGACGATCNNWNNW")

    # Leave a blank row after writing the 'ForwardPrimer(Fp)' data
    ws_dna.cell(20 + len(df_sample['ForwardPrimer(Fp)']), 4).value = None

    # Write the 'ReversePrimer(Rp)' data to the sheet
    write_to_sheet(df_sample['ReversePrimer(Rp)'], 4, 21 + len(df_sample['ForwardPrimer(Fp)']),
                   "CTTGGCACCCGAGAATTCCANNWNNW")
    ws_trial = wb['实验用引物对（不需要订购合成）']

    row_index = 1
    # Iterate through the rows of the dataframe
    for r in dataframe.iterrows():
        # Get the row data
        row_data = r[1]
        # Get the cell values
        values = row_data.tolist()
        # Write the data to the sheet
        ws_trial.append(values)
        # Increment the row index
        row_index += 1

    # Save the workbook
    save_file = os.path.join(order_path, '{}_{}_{}.xlsx'.format(sampleID, os.path.basename(
        (config['order_template']['sh']).split('.')[0]), datetime.datetime.now().strftime("%Y%m%d%H%M%S")))
    wb.save(save_file)
    return save_file


def write_hz_order(df_sample, dataframe, order_path, sampleID):
    os.makedirs(order_path, exist_ok=True)
    try:
        order_template = config['order_template']['hz']
        wb = openpyxl.load_workbook(order_template)
    except Exception as e:
        logging.error(f'ERROR: {e}')
        sys.exit(1)
    ws_dna = wb['DNA合成订购表']

    # Define a function to write the data to the sheet
    def write_to_sheet(data, column, start_row, sequence_prefix=None):
        for i, value in enumerate(data):
            if value.islower():
                value = value.upper()
            if sequence_prefix:
                value = sequence_prefix + value
            ws_dna.cell(start_row + i, column).value = value
            ws_dna.cell(start_row + i, 5).value = 2
            ws_dna.cell(start_row + i, 7).value = 'PAGE'
            ws_dna.cell(start_row + i, 8).value = 2
            ws_dna.cell(start_row + i, 11).value = '1管TE溶解为50uM浓度，1管干粉'

    # Write the 'F_id' data to the sheet
    write_to_sheet(df_sample['F_id'], 2, 20)

    # Leave a blank row after writing the 'F_id' data
    ws_dna.cell(20 + len(df_sample['F_id']), 2).value = None

    # Write the 'R_id' data to the sheet
    write_to_sheet(df_sample['R_id'], 2, 21 + len(df_sample['F_id']))

    # Write the 'ForwardPrimer(Fp)' data to the sheet
    write_to_sheet(df_sample['ForwardPrimer(Fp)'], 3, 20, "GTTCAGAGTTCTACAGTCCGACGATCNNWNNW")

    # Leave a blank row after writing the 'ForwardPrimer(Fp)' data
    ws_dna.cell(20 + len(df_sample['ForwardPrimer(Fp)']), 3).value = None

    # Write the 'ReversePrimer(Rp)' data to the sheet
    write_to_sheet(df_sample['ReversePrimer(Rp)'], 3, 21 + len(df_sample['ForwardPrimer(Fp)']),
                   "CTTGGCACCCGAGAATTCCANNWNNW")

    ws_trial = wb['实验用引物对（不需要订购合成）']

    row_index = 1
    # Iterate through the rows of the dataframe
    for r in dataframe.iterrows():
        # Get the row data
        row_data = r[1]
        # Get the cell values
        values = row_data.tolist()
        # Write the data to the sheet
        ws_trial.append(values)
        # Increment the row index
        row_index += 1

    # Save the workbook
    save_file = os.path.join(order_path, '{}_{}_{}.xlsx'.format(sampleID, os.path.basename(
        (config['order_template']['hz']).split('.')[0]), datetime.datetime.now().strftime("%Y%m%d%H%M%S")))
    wb.save(save_file)
    return save_file


def write_dg_order(df_sample, dataframe, order_path, sampleID):
    os.makedirs(order_path, exist_ok=True)
    try:
        order_template = config['order_template']['dg']
        wb = openpyxl.load_workbook(order_template)
    except Exception as e:
        logging.error(f'ERROR: {e}')
        sys.exit(1)
    ws_dna = wb['订单表格']

    # Define a function to write the data to the sheet
    def write_to_sheet(data, column, start_row, sequence_prefix=None):
        for i, value in enumerate(data):
            if value.islower():
                value = value.upper()
            if sequence_prefix:
                value = sequence_prefix + value
            ws_dna.cell(start_row + i, column).value = value
            # ws_dna.cell(start_row + i, 5).value = 2
            ws_dna.cell(start_row + i, 8).value = 'PAGE'
            ws_dna.cell(start_row + i, 14).value = 2
            ws_dna.cell(start_row + i, 13).value = '1管TE溶解为50uM浓度，1管干粉'

    # Write the 'F_id' data to the sheet
    write_to_sheet(df_sample['F_id'], 5, 16)

    # Write the 'R_id' data to the sheet
    write_to_sheet(df_sample['R_id'], 5, 16 + len(df_sample['F_id']))

    # Write the 'ForwardPrimer(Fp)' data to the sheet
    write_to_sheet(df_sample['ForwardPrimer(Fp)'], 6, 16, "GTTCAGAGTTCTACAGTCCGACGATCNNWNNW")

    # Write the 'ReversePrimer(Rp)' data to the sheet
    write_to_sheet(df_sample['ReversePrimer(Rp)'], 6, 16 + len(df_sample['ForwardPrimer(Fp)']),
                   "CTTGGCACCCGAGAATTCCANNWNNW")

    ws_trial = wb['实验用引物对（不需要订购合成）']
    row_index = 1
    # Iterate through the rows of the dataframe
    for r in dataframe.iterrows():
        # Get the row data
        row_data = r[1]
        # Get the cell values
        values = row_data.tolist()
        # Write the data to the sheet
        ws_trial.append(values)
        # Increment the row index
        row_index += 1
    # Save the workbook
    save_file = os.path.join(order_path, '{}_{}_{}.xlsx'.format(sampleID, os.path.basename(
        (config['order_template']['dg']).split('.')[0]), datetime.datetime.now().strftime("%Y%m%d%H%M%S")))
    wb.save(save_file)
    return save_file


def write_sg_order(df_order, df_combined, order_dir, sampleID):
    """
    Writes primer order data to an Excel workbook based on a template and saves it.

    :param df_order: DataFrame containing the primer order data to be written to the '引物合成订购表' sheet.
    :param df_combined: DataFrame containing additional data to be written to the '实验用引物对（不需要订购合成）' sheet.
    :param order_dir: Directory where the output Excel file will be saved.
    :param sampleID: Sample identifier used as part of the output file name.
    :return: The path of the saved Excel workbook.
    """
    try:
        order_template = config['order_template']['sg']
        wb = openpyxl.load_workbook(order_template)
    except Exception as e:
        raise e

    ws_dna = wb['引物合成订购表']

    def write_to_sheet(dataframe, start_row):
        """ Helper function to write data to the Excel sheet """
        for i, raw in dataframe.iterrows():
            ws_dna.cell(start_row + i, 2, raw['PrimerName'])
            ws_dna.cell(start_row + i, 3, raw['Sequence'])
            ws_dna.cell(start_row + i, 5, raw['TubeCount'])
            ws_dna.cell(start_row + i, 7, raw['PurificationMethod'])
            ws_dna.cell(start_row + i, 8, raw['Nmoles'])
            ws_dna.cell(start_row + i, 12, raw['Remarks'])

    write_to_sheet(df_order, 18)

    ws_trial = wb['实验用引物对（不需要订购合成）']
    ws_trial.append(df_combined.columns.tolist())  # Write the column names
    for index, row in df_combined.iterrows():
        ws_trial.append(row.tolist())  # Write the data

    save_file = os.path.join(order_dir,
                             f'{sampleID}_{os.path.basename(order_template).split(".")[0]}_{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx')
    wb.save(save_file)

    return save_file


def upsert_to_database(df, table_name, unique_col, update_cols):
    """
    Update specific fields in the database based on the unique_col or insert a new record.

    :param df: DataFrame containing the data to update or insert.
    :param table_name: The name of the table to upsert into.
    :param unique_col: The name of the column to match for the update.
    :param update_cols: List of column names that need to be updated.
    """
    engine = db_handler.get_engine()
    inspector = db_handler.get_inspector()

    # Check if the table exists
    if not inspector.has_table(table_name):
        save_to_database(df, table_name)
        return

    # Start a transaction
    with engine.begin() as conn:
        for index, row in df.iterrows():
            try:
                # Check if a record with the unique column value exists
                exists_stmt = text(f"""SELECT EXISTS (
                    SELECT 1 FROM {table_name} WHERE {unique_col} = :value
                )""")
                exists_result = conn.execute(exists_stmt, {'value': row[unique_col]}).scalar()

                if exists_result:
                    # Record exists, construct an update statement
                    update_values = {col: row[col] for col in update_cols}
                    update_values['unique_value'] = row[unique_col]
                    update_stmt = text(f"""
                            UPDATE {table_name} SET 
                            {', '.join([f"{col} = :{col}" for col in update_cols])} 
                            WHERE {unique_col} = :unique_value
                        """)
                    conn.execute(update_stmt, update_values)
                    logger.info(f"Updated record with {unique_col} = {row[unique_col]} in '{table_name}' table.")
                else:
                    # Record does not exist, insert the new record
                    save_to_database(df.iloc[[index]], table_name)
                    logger.info(
                        f"Inserted new record with {unique_col} = {row[unique_col]} into '{table_name}' table.")
            except SQLAlchemyError as e:
                logger.error(f"An error occurred: {e}")


def write_order(sampleID, df_design, df_res, order_dir, mold, skip_snp_design, send_email=True):
    """
    Processes primer design results, prepares the primer order, writes the order file, and updates the database.

    :param sampleID: The unique identifier of the sample.
    :param df_design: DataFrame containing the design information.
    :param df_res: DataFrame containing the primer design results.
    :param order_dir: Directory where the order file will be saved.
    :param mold: String representing the mold (e.g., 'sh', 'hz', 'sg', 'dg') for different order templates.
    :param skip_snp_design: Boolean flag indicating whether to skip SNP design.
    :param send_email: Boolean flag indicating whether to send an email for quality control. Default is True.

    :return: The path to the written order file.

    This function takes the results of primer design, processes them for quality control, and prepares the order
    file based on the specified mold. It then logs the order information in a database for tracking.
    """
    # Check the primer results
    df_sample = process_primer_results(df_res, df_design, sampleID, skip_snp_design, send_email=send_email)

    # Add order information and merge primers
    df_processed = process_primer_sample(df_sample)

    # Added order format information
    df_order = process_primer_order(df_processed, mold)

    # The mold value is mapped to the corresponding function
    order_functions = {
        'sh': write_sh_order,
        'hz': write_hz_order,
        'sg': write_sg_order,
        'dg': write_dg_order
    }

    # appropriate function is called and logged
    if mold in order_functions:
        primer_result = order_functions[mold](df_order, df_processed, order_dir, sampleID)
        logger.info(f'Primer design {mold.upper()} order template writing completed.')
    else:
        logger.error(f'Unknown mold: {mold}')
        sys.exit(1)

    df_order_info = pd.DataFrame({
        "SampleID": [sampleID],  # 样本ID
        "OrderFile": [primer_result],  # 订单文件
        "ReviewStatus": [""],  # 审核状态
        "EmailSent": [0],  # 是否发送过邮件：0 - 未发送订购单，1 - 已发送过订购单，2 - 不需要发送订购
        "DesignDate": [pd.to_datetime(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))],  # 设计完成时间
        "OrderDate": [datetime.date.today()],  # 订购时间
        "OrderCompany": [mold]  # 订购公司
    }, index=[0])

    # Write or update order history
    upsert_to_database(df_order_info, 'monitor_order', 'SampleID', ['OrderFile', 'DesignDate'])

    return primer_result


def check_email_sent(sample_id, table_name):
    """
    Checks the email sent status for a given sample ID.

    :param sample_id: The sample ID to check in the database.
    :param table_name: The name of the table in the database.
    :return: An integer indicating the status of the email sent.
             Returns 0 if an email has not been sent, 1 if sent, 2 if not required,
             and None if no record is found or for any other unexpected value.
    """
    engine = db_handler.get_engine()

    # Construct the SQL query using text()
    query = text(f"""
        SELECT EmailSent FROM {table_name}
        WHERE SampleID = :sample_id
    """)

    # Execute the query and fetch the result
    with engine.connect() as connection:
        result = connection.execute(query, {'sample_id': sample_id}).fetchone()

        # Check if the sample ID was found and return the appropriate value
        if result:
            # If result is a tuple, the EmailSent column is assumed to be at index 0
            email_sent_status = result[0]
            if email_sent_status in [0, 1, 2]:
                return email_sent_status
            else:
                return None  # This covers empty or any other unexpected value
        else:
            # If no result is found for the sample ID, also return None
            return None


def update_email_status(sample_id, table_name, review_status=None, email_sent=1):
    """
    Updates the database to mark the email sent status for a given sample ID, and optionally update the review status.

    :param sample_id: The sample ID in the database.
    :param table_name: The name of the table in the database.
    :param review_status: Optional. The new review status for the sample.
    :param email_sent: The status to set for EmailSent. Default is 1.
    """
    engine = db_handler.get_engine()

    # Prepare the SQL statement to update the EmailSent, OrderDate, and optionally ReviewStatus
    update_values = "EmailSent = :email_sent, OrderDate = :order_date"
    if review_status is not None:
        update_values += ", ReviewStatus = :review_status"

    update_stmt = text(f"""
        UPDATE {table_name}
        SET {update_values}
        WHERE SampleID = :sample_id
    """)

    # Prepare a reference dictionary
    order_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    params = {
        "email_sent": email_sent,
        "order_date": order_date,
        "sample_id": sample_id
    }

    if review_status is not None:
        params["review_status"] = review_status

    # Execute the update
    with engine.begin() as conn:
        conn.execute(update_stmt, params)
        update_msg = f"EmailSent updated to {email_sent} and OrderDate set to {order_date}"
        if review_status:
            update_msg += f", with ReviewStatus updated to {review_status}"
        update_msg += f" for SampleID: {sample_id}"
        logger.info(update_msg)


def generate_testing_periods(cycle, max_cycle):
    periods = [cycle]
    for i in range(cycle * 2, max_cycle + 1, cycle):
        periods.append(i)
    if periods[-1] != max_cycle:
        periods.append(max_cycle)
    return periods


def check_order(sampleID, primer_result, skip_review, send_email=True):
    """
    Monitors and manages the process of checking sample audit status, sending emails, and updating database.

    :param sampleID: The unique identifier for the sample.
    :param primer_result: The file path of the primer result that will be attached to the email.
    :param skip_review: A flag to determine whether to skip the review process. If True, the email is sent immediately.
    :param send_email: Boolean flag indicating whether to send an email for quality control. Default is True.

    The function implements a while loop that continuously checks the audit status of the sample.
    If the sample passes the audit, an email is sent, and the database is updated.
    If the sample is still under review, the function waits for a predefined interval before rechecking.
    If the sample fails the audit or experiences an anomaly, an alert email is sent, and the function exits.
    """
    check_toaddrs = config['emails']['setup']['log_toaddrs']
    qc_toaddrs = config['emails']['setup']['qc_toaddrs']
    order_toaddrs = config['emails']['setup']['order_toaddrs']
    log_cc = config['emails']['setup']['log_cc']
    order_cc = config['emails']['setup']['cc']
    test_subject = f'样本引物合成订购 ( 预先订购 | 位点追加 | 项目测试 | 科研项目) - {sampleID} '
    pro_subject = f'样本引物合成订购 (自动发送) - {sampleID} '
    test_message = f'样本ID：{sampleID}\n注：该引物合成用于(预先订购 | 位点追加 | 项目测试 | 科研项目)其中之一。\n引物结果：{os.path.basename(primer_result)}（见附件）'
    pro_message = f'样本ID：{sampleID}\nCMS审核结果：已通过\n引物结果：{os.path.basename(primer_result)}（见附件）'

    toaddrs = check_toaddrs if DEBUG else order_toaddrs
    qc_toaddrs = check_toaddrs if DEBUG else qc_toaddrs
    cc = log_cc if DEBUG else order_cc
    subject = test_subject if skip_review else pro_subject
    message = test_message if skip_review else pro_message

    email_status = check_email_sent(sampleID, 'monitor_order')

    if email_status == 0:
        # 如果EmailSent是0，发送邮件并更新数据库
        if skip_review:
            if send_email:
                emit(subject, message, attachments=[primer_result], to_addrs=toaddrs, cc_addrs=cc)
                update_email_status(sampleID, 'monitor_order')

        else:
            program_time = datetime.datetime.now()
            start_time = datetime.datetime.now()
            check_interval_minutes = int(config['check_interval_minutes'])  # 检测时间
            check_frequency = datetime.timedelta(minutes=check_interval_minutes)  # min转换s
            email_cycle = int(config['email_interval_days'])  # 邮件预警周期
            max_days_to_check = int(config['max_interval_days'])  # 最大检测周期
            email_days = generate_testing_periods(email_cycle, max_days_to_check)  # 周期内天数
            is_first_check = True  # 持续监测预警
            last_email_sent = None  # 发送邮件标志

            while True:
                status_abbr, status_desc = get_audit_status(sampleID)
                review_status = f'{status_abbr}({status_desc})'

                # 审核通过
                if status_abbr in ['YWC', 'YSH', 'BGYSH']:
                    if send_email:
                        emit(subject, message, attachments=[primer_result], to_addrs=toaddrs, cc_addrs=cc)
                        update_email_status(sampleID, 'monitor_order', review_status)
                    logger.info('Complete the sample primer design and send the order!')
                    break

                # 检测中
                elif status_abbr in ['JCZ', 'DSH', 'BGDSH']:
                    if is_first_check:
                        subject = f'样本审核状态持续检测 - {sampleID}'
                        message = f'样本ID {sampleID} 审核状态持续检测中···\n目前样本审核状态：{review_status}。\n注意：在订单发送之前，审核人员可查看附件的引物订单检查错误，并告知程序管理人员终止自动发送程序。\n提示：程序会按照自定义时间检测CMS系统审核状态，等待审核状态发生改变，该引物订单会自动发送订购。'
                        if send_email:
                            emit(subject, message, attachments=[primer_result], to_addrs=qc_toaddrs)
                        is_first_check = False

                    # 根据自定义时间间隔等待下一次检测
                    current_time = datetime.datetime.now()

                    # 等待时间
                    time_to_wait = check_frequency - (current_time - start_time)
                    if time_to_wait.total_seconds() > 0:
                        time.sleep(time_to_wait.total_seconds())

                    # 等待后继续算时间间隔
                    current_time = datetime.datetime.now()
                    time_diff = current_time - start_time
                    logger.info(f'The current time is {current_time}, Waiting {time_diff}')

                    # 若时间间隔已经大于检测间隔时间
                    if time_diff >= check_frequency:
                        # 计算过去的天数
                        days_since_last = (current_time - program_time).days
                        days_since_last_email = int(days_since_last)
                        program_time_formatted = program_time.strftime('%Y-%m-%d %H:%M:%S:%f')
                        current_time_formatted = current_time.strftime('%Y-%m-%d %H:%M:%S:%f')
                        if days_since_last_email in email_days:
                            if days_since_last_email < max_days_to_check:
                                if last_email_sent is None or (current_time - last_email_sent).days >= email_cycle:
                                    subject = f'样本审核状态超过 {days_since_last_email} 天未更新提醒 - {sampleID} '
                                    message = f'样本ID：{sampleID}\nCMS审核结果：检测到已经超过 {days_since_last_email} 天未通过审核，请审核人员检查并更新状态！\n检测时间：{program_time_formatted} —— {current_time_formatted}\n 。'
                                    if send_email:
                                        emit(subject, message, to_addrs=qc_toaddrs)
                                    last_email_sent = current_time
                            else:
                                if last_email_sent is None or (current_time - last_email_sent).days >= email_cycle:
                                    subject = f'样本审核状态超过半个月未更新警告 - {sampleID} '
                                    message = f'样本ID：{sampleID}\nCMS审核结果：检测到已经超过半个月未通过审核，请审核人员检查并更新状态！\n检测时间：{program_time_formatted} —— {current_time_formatted}\n警告：该样本审核状态最后一次检测，程序将自动退出以防止进一步的数据处理。\n请立即检查相关数据并采取适当措施。'
                                    if send_email:
                                        emit(subject, message, to_addrs=qc_toaddrs)
                                    logger.info(
                                        f'The program has been running for more than {max_days_to_check} days. Exiting program.')
                                    break
                        # 重置 start_time
                        start_time = datetime.datetime.now()

                    else:
                        time_to_wait = check_frequency - time_diff
                        if time_to_wait.total_seconds() > 0:
                            time.sleep(time_to_wait.total_seconds())

                # 检测终止
                else:
                    subject = f'样本状态检测异常警告 - {sampleID}'
                    message = f'警告：样本ID {sampleID} 样本状态检测异常。\nCMS审核状态：{review_status}\n提示：程序将自动退出以防止进一步的数据处理。\n请立即检查相关数据并采取适当措施。'
                    if send_email:
                        emit(subject, message, to_addrs=qc_toaddrs)
                        update_email_status(sampleID, 'monitor_order', review_status=review_status,
                                            email_sent=2)
                    logger.error(
                        f'Sample ID {sampleID} Detect the anomaly and exit the program. Please check the relevant data immediately and take appropriate action.')
                    sys.exit(1)

    elif email_status in [1, 2]:
        # 如果EmailSent是1，已经发送过邮件了，如果DEBUG为True还会发送至测试邮箱
        # 如果EmailSent是2，不发送邮件，项目终止了
        if email_status == 1 and DEBUG:
            if send_email:
                emit(subject, message, attachments=[primer_result], to_addrs=qc_toaddrs)
                logger.info(f"Email sent to DEBUG addresses for SampleID: {sampleID}.")
        else:
            logger.info(f"No action needed for SampleID: {sampleID} as EmailSent is {email_status}")
        sys.exit(0)

    else:
        # 如果EmailSent是None或其他值，发送错误消息并退出程序
        logger.error(f"Unexpected EmailSent status for SampleID: {sampleID}")
        sys.exit(1)


def execute(args):
    global DEBUG
    global db_handler

    # 确定 DEBUG 模式：如果命令行参数指定了 --debug，则使用该参数，否则使用配置文件中的设置
    DEBUG = args.debug if args.debug else config.get('DEBUG', False)

    # 使用 DEBUG 变量
    if DEBUG:
        # 运行调试模式下的代码
        logger.info("Running in debug mode...")
    else:
        # 运行非调试模式下的代码
        logger.info("Running in normal mode...")

    # 连接数据库
    db_config = config['DB_CONFIG']
    db_url = f"mysql+pymysql://{db_config['user']}:{db_config['passwd']}@{db_config['host']}:{db_config['port']}/{db_config['db']}"
    db_handler = pt.DatabaseHandler(db_url)

    # 设置参数变量
    mold = args.mold
    file_path = args.input_file
    output_dir = args.output_dir
    url = args.url
    send_email = args.send_email
    cancer_id = args.cancer_id
    email_interval = args.email_interval
    exit_threshold = args.exit_threshold
    no_timeout = args.no_timeout
    skip_snp_design = args.skip_snp
    skip_hot_design = args.skip_hot
    skip_driver_design = args.skip_driver
    skip_check = args.skip_check
    skip_review = args.skip_review
    run_order = args.run_order

    # 输出文件夹处理
    outcome_dir = os.path.join(os.path.abspath(output_dir), 'primer_outcome')
    os.makedirs(outcome_dir, exist_ok=True)
    order_dir = os.path.join(os.path.abspath(output_dir), 'primer_order')
    os.makedirs(order_dir, exist_ok=True)

    # 获取样本ID
    sampleID = get_sample_id(file_path)

    # 检查日期，默认10天发邮件提示，超过30天退出程序
    if not no_timeout:
        check_sample_date(sampleID, send_email=send_email, email_interval=email_interval, exit_threshold=exit_threshold)

    # 是否跳过样本检查
    if not skip_check:
        # 检查样本项目
        handle_mrd_sample(sampleID, send_email=send_email)

    # 读取选点文件
    df = read_loci_file(file_path)

    # 判断和处理选点
    df_loci = loci_examined(df, skip_snp_design, skip_hot_design, skip_driver_design, cancer_id=cancer_id,
                            send_email=send_email)

    # 添加 templateID
    df_design = add_templateID(df_loci)

    # 优先 driver 基因进行引物设计
    df_no_driver, design_num, driver_list, driver_str = process_driver(df_design, url, outcome_dir, sampleID,
                                                                       skip_driver_design)

    # 循环设计引物
    df_res = perform_primer_design(df_no_driver, sampleID, url, outcome_dir, design_num, driver_list, driver_str)

    # 写入订单表
    primer_result = write_order(sampleID, df_design, df_res, order_dir, mold, skip_snp_design, send_email=send_email)

    # 检查订单状态
    if run_order:
        check_order(sampleID, primer_result, skip_review, send_email)

    # 返回订单表
    print(primer_result)


def main():
    # 设置命令行参数
    parser = argparse.ArgumentParser(description='Automatic primer design.')

    # 必需的参数
    parser.add_argument('-m', '--mold', required=True, dest='mold', choices=['sh', 'hz', 'sg', 'dg'],
                        help='Currently, the order template is only available in sh(上海百力格), hz(湖州河马), sg(上海生工), dg(上海迪赢).')
    parser.add_argument('-i', '--input_file', required=True, dest='input_file',
                        help='Input file path for primer design.')
    parser.add_argument('-o', '--output_dir', required=True, dest='output_dir',
                        help='Output directory for primer results and orders.')

    # 可选参数
    parser.add_argument('--url', default=config['mfe_primer'], dest='url',
                        help='URL for primer design API.')
    parser.add_argument('--no-email', action='store_false', default=True, dest='send_email',
                        help='Do not send email if set.')
    parser.add_argument('--c-id', dest='cancer_id',
                        help='Cancer ID, if applicable.')
    parser.add_argument('--email-freq', type=int, default=10, dest='email_interval',
                        help='Frequency in days for sending reminder emails.')
    parser.add_argument('--exit-lim', type=int, default=30, dest='exit_threshold',
                        help='Time limit in days to stop checking and exit.')
    parser.add_argument('--no-timeout', action='store_true', dest='no_timeout',
                        help='Disable time-based program exit.')
    parser.add_argument('--skip-snp', action='store_true', dest='skip_snp',
                        help='Skip SNP design if set.')
    parser.add_argument('--skip-hot', action='store_true', dest='skip_hot',
                        help='Skip hot design if set.')
    parser.add_argument('--skip-driver', action='store_true', dest='skip_driver',
                        help='Skip driver design if set.')
    parser.add_argument('--skip-check', action='store_true', dest='skip_check',
                        help='Skip system check if set.')
    parser.add_argument('--skip-review', action='store_true', dest='skip_review',
                        help='Skip review process if set.')
    parser.add_argument('--run-order', action='store_true', dest='run_order',
                        help='Run the check_order function if set.')
    parser.add_argument('--debug', action='store_true', dest='debug',
                        help='Run in debug mode.')

    # 解析命令行参数
    args = parser.parse_args()
    execute(args)


if __name__ == '__main__':
    main()
